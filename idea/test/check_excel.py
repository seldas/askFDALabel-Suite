from openpyxl import load_workbook

try:
    wb = load_workbook('tmp/import_example_fdalabel.xlsx', data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if rows:
        print("Columns:", rows[0])
        if len(rows) > 1:
            print("First row:", dict(zip(rows[0], rows[1])))
except Exception as e:
    print("Error:", e)