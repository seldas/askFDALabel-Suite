from flask import Blueprint, render_template, request, redirect, url_for, jsonify, send_from_directory, current_app
from flask_login import login_required, current_user
import os
import zipfile
import re
import json
from difflib import HtmlDiff

from dashboard.srcs.extensions import db
from dashboard.srcs.models import Favorite, Annotation, FavoriteComparison
from dashboard.srcs.services.xml_handler import (
    parse_spl_xml,
    extract_metadata_from_xml,
    flatten_sections,
    get_aggregate_content
)
from dashboard.srcs.services.fda_client import (
    get_label_metadata,
    get_label_xml,
    find_labels,
    find_labels_by_set_ids
)
from dashboard.srcs.utils import (
    normalize_text_for_diff,
    extract_numeric_section_id,
    normalize_title_text,
    get_section_sort_key
)
from dashboard.srcs.config import Config
from dashboard.srcs.services.fdalabel_db import FDALabelDBService

main_bp = Blueprint('main', __name__)

@main_bp.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(current_app.root_path, 'static'),
                               'favicon.svg', mimetype='image/svg+xml')

@main_bp.route('/')
def index():
    """ Redirects to the Next.js dashboard. """
    return redirect('/dashboard')


@main_bp.route('/info')
def info():
    """ Renders the information page about the tool. """
    return render_template('info.html')


@main_bp.route('/upload_label', methods=['POST'])
def upload_label():
    """
    Handles uploading of an XML or ZIP file containing an SPL label.
    Extracts, validates (PLR/Non-PLR), and stores the label for comparison.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    filename = file.filename.lower()
    if not (filename.endswith('.xml') or filename.endswith('.zip')):
        return jsonify({'error': 'Invalid file format. Please upload .xml or .zip.'}), 400

    # Read current set_ids to validate against
    current_set_ids = request.form.getlist('current_set_ids[]') # AJAX sends array as []
    
    xml_content = None
    
    try:
        if filename.endswith('.zip'):
            with zipfile.ZipFile(file) as z:
                # Find first XML file
                xml_files = [f for f in z.namelist() if f.lower().endswith('.xml')]
                if not xml_files:
                    return jsonify({'error': 'No XML file found in the ZIP archive.'}), 400
                
                with z.open(xml_files[0]) as f:
                    xml_content = f.read().decode('utf-8')
        else:
            xml_content = file.read().decode('utf-8')
            
        if not xml_content:
             return jsonify({'error': 'Empty file content.'}), 400

        # Extract metadata
        meta = extract_metadata_from_xml(xml_content)
        if not meta or not meta.get('set_id'):
            return jsonify({'error': 'Could not parse SPL metadata or Set ID from the XML.'}), 400

        new_label_format = meta.get('label_format')
        new_set_id = meta.get('set_id')

        # VALIDATION: Check against existing labels
        if current_set_ids:
            # We need to check the format of at least one existing label
            # We can use get_label_metadata which now checks local first
            existing_meta = get_label_metadata(current_set_ids[0])
            if existing_meta:
                existing_format = existing_meta.get('label_format')
                if existing_format and new_label_format and existing_format != new_label_format:
                     return jsonify({
                         'error': f"Format mismatch: The uploaded label is '{new_label_format}', but you are comparing '{existing_format}' labels. Please upload a compatible label."
                     }), 400

        # Save to local storage
        save_path = os.path.join(Config.UPLOAD_FOLDER, f"{new_set_id}.xml")
        with open(save_path, 'w', encoding='utf-8') as f:
            f.write(xml_content)
            
        return jsonify({'success': True, 'set_id': new_set_id})

    except Exception as e:
        current_app.logger.error(f"Error processing upload: {e}")
        return jsonify({'error': f"Error processing file: {str(e)}"}), 500

import uuid
from openpyxl import load_workbook

@main_bp.route('/import_fdalabel', methods=['POST'])
def import_fdalabel():
    """
    Handles importing drug labels from an FDALabel Excel file.
    Parses the Excel and redirects to the selection page with the data.
    """
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No selected file'}), 400

    if not file.filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'error': 'Invalid file format. Please upload .xlsx'}), 400

    try:
        wb = load_workbook(file, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            return jsonify({'success': False, 'error': 'The Excel file is empty.'}), 400
            
        columns = rows[0]
        # Create a mapping of column name to index
        col_map = {name: i for i, name in enumerate(columns) if name}
        
        required_cols = ['SET ID', 'Trade Name']
        missing = [c for c in required_cols if c not in col_map]
        if missing:
            return jsonify({'success': False, 'error': f'Missing required columns: {", ".join(missing)}'}), 400

        # Helper to safely get value
        def get_val(row, col_name):
            idx = col_map.get(col_name)
            if idx is not None and idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return 'N/A'

        labels = []
        for row in rows[1:]:
            if not row[col_map['SET ID']]:
                continue
                
            set_id = str(row[col_map['SET ID']]).strip()
            
            # Map Excel columns to our label structure
            brand_name = get_val(row, 'Trade Name')
            generic_name = get_val(row, 'Generic/Proper Name(s)')
            company = get_val(row, 'Company')
            effective_date = get_val(row, 'SPL Effective Date (YYYY/MM/DD)')
            app_num = get_val(row, 'Application Number(s)')
            labeling_type = get_val(row, 'Labeling Type')
            ndcs = get_val(row, 'NDC(s)')
            
            # Additional FDALabel specific columns
            dosage_forms = get_val(row, 'Dosage Form(s)')
            routes = get_val(row, 'Route(s) of Administration')
            marketing_category = get_val(row, 'Marketing Category')
            epc = get_val(row, 'Established Pharmacologic Class(es)')
            initial_approval = get_val(row, 'Initial U.S. Approval')
            marketing_dates = get_val(row, 'Marketing Date(s) (YYYY/MM/DD)')
            active_ingredient_unii = get_val(row, 'Active Ingredient UNII(s)')
            active_ingredients = get_val(row, 'Active Ingredient(s)')
            active_moiety_name = get_val(row, 'Active Moiety Name(s)')
            active_moiety_unii = get_val(row, 'Active Moiety UNII(s)')
            fdalabel_link = get_val(row, 'FDALabel Link')
            dailymed_spl_link = get_val(row, 'DailyMed SPL Link')
            dailymed_pdf_link = get_val(row, 'DailyMed PDF Link')

            # Determine format - removed as requested, using Marketing Category instead
            label_format = None
            
            # Simplified product type
            prod_type = 'Rx' if 'PRESCRIPTION' in labeling_type.upper() else 'OTC'
            if 'OTC' in labeling_type.upper(): prod_type = 'OTC'

            labels.append({
                'set_id': set_id,
                'brand_name': brand_name,
                'generic_name': generic_name,
                'manufacturer_name': company,
                'effective_time': effective_date,
                'label_format': label_format,
                'labeling_type': labeling_type,
                'application_number': app_num,
                'product_type': prod_type,
                'ndc': ndcs,
                'marketing_category': marketing_category,
                'dosage_forms': dosage_forms,
                'routes': routes,
                'epc': epc,
                'initial_approval': initial_approval,
                'marketing_dates': marketing_dates,
                'active_ingredient_unii': active_ingredient_unii,
                'active_ingredients': active_ingredients,
                'active_moiety_name': active_moiety_name,
                'active_moiety_unii': active_moiety_unii,
                'fdalabel_link': fdalabel_link,
                'dailymed_spl_link': dailymed_spl_link,
                'dailymed_pdf_link': dailymed_pdf_link,
                'source': 'excel'
            })

        if not labels:
            return jsonify({'success': False, 'error': 'No valid labels found in the Excel file.'}), 400

        # Store in temporary file
        import_id = str(uuid.uuid4())
        import_filename = f"import_{import_id}.json"
        import_path = os.path.join(Config.UPLOAD_FOLDER, import_filename)
        
        with open(import_path, 'w', encoding='utf-8') as f:
            json.dump(labels, f)

        return jsonify({
            'success': True, 
            'redirect_url': url_for('main.search', import_id=import_id)
        })

    except Exception as e:
        current_app.logger.error(f"Error importing Excel: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@main_bp.route('/search', methods=['GET', 'POST'])
def search():
    """ Handles drug name search with pagination. """
    if request.method == 'POST':
        drug_name = request.form.get('drug_name')
        if not drug_name:
            return "Please enter a drug name.", 400
        # Redirect to a GET request to make the URL bookmarkable
        return redirect(url_for('main.search', drug_name=drug_name))

    drug_name = request.args.get('drug_name')
    batch_id_search = request.args.get('batch_id_search')
    import_id = request.args.get('import_id')

    if not drug_name and not batch_id_search and not import_id:
        return redirect(url_for('main.index'))

    page = request.args.get('page', 1, type=int)
    view = request.args.get('view')
    
    # Limitation Logic: OpenFDA = 10, FDALabel/Internal = 100000
    from dashboard.srcs.services.fdalabel_db import FDALabelDBService
    if FDALabelDBService.check_connectivity() or import_id:
        limit = 100000
    else:
        limit = 10 # OpenFDA limit
        
    skip = (page - 1) * limit

    if import_id:
        import_path = os.path.join(Config.UPLOAD_FOLDER, f"import_{import_id}.json")
        if os.path.exists(import_path):
            with open(import_path, 'r', encoding='utf-8') as f:
                all_imported_labels = json.load(f)
            total = len(all_imported_labels)
            labels = all_imported_labels[skip : skip + limit]
            drug_name_display = "Excel Import"
            page_title = "Imported FDALabel Results"
        else:
            return "Import session expired or not found.", 404
    elif batch_id_search:
        ids_list = [sid.strip() for sid in batch_id_search.split(',') if sid.strip()]
        labels, total = find_labels_by_set_ids(ids_list, skip=skip, limit=limit)
        drug_name_display = "Identifier Batch Search"
        page_title = "Batch Search Results"
    else:
        labels, total = find_labels(drug_name, skip=skip, limit=limit)
        drug_name_display = drug_name
        page_title = f"Search Results for \"{drug_name}\""

    # Determine default view if not specified
    if not request.args.get('view'):
        if total > 5 or batch_id_search or import_id:
             view = 'table'
        else:
             view = 'panel'

    # Server-side favorite check removed to prevent cross-project confusion. 
    # Client-side JS will update status based on active project.
    # if labels and current_user.is_authenticated:
    #     user_favs = Favorite.query.filter_by(user_id=current_user.id).all()
    #     user_favorite_ids = {f.set_id for f in user_favs}
    #     for label in labels:
    #         label['is_favorite'] = label['set_id'] in user_favorite_ids

    if not labels and page == 1:
        # No results found
        if request.headers.get('Accept') == 'application/json' or request.args.get('json') == '1':
            return jsonify({
                "drug_name": drug_name_display,
                "page_title": page_title,
                "labels": [],
                "total": 0,
                "page": page,
                "limit": limit,
                "view": view
            })
        return render_template('selection.html', drug_name=drug_name_display, page_title=page_title, labels=None, total=0, page=page, limit=limit, view=view)

    # If multiple results (or a batch search, or single result), show the selection page
    is_internal = FDALabelDBService.check_connectivity()
    
    if request.headers.get('Accept') == 'application/json' or request.args.get('json') == '1':
        return jsonify({
            "drug_name": drug_name_display,
            "page_title": page_title,
            "labels": labels,
            "total": total,
            "page": page,
            "limit": limit,
            "view": view,
            "is_internal": is_internal,
            "batch_id_search": batch_id_search,
            "import_id": import_id
        })

    return render_template('selection.html', drug_name=drug_name_display, page_title=page_title, labels=labels, total=total, page=page, limit=limit, view=view, is_internal=is_internal)


@main_bp.route('/label/<set_id>')
def view_label(set_id):
    """ Displays the parsed SPL label for a given set_id. """
    drug_name_from_query = request.args.get('drug_name', '')
    import_id = request.args.get('import_id')
    
    label_xml_raw = get_label_xml(set_id)
    if not label_xml_raw:
        return "Could not fetch the label from DailyMed.", 500

    doc_title, sections, fallback_html, highlights, table_of_contents = parse_spl_xml(label_xml_raw, set_id)
    
    # Use the drug name from the query for the title, but fall back to the parsed title
    display_drug_name = drug_name_from_query if drug_name_from_query else doc_title

    # Load saved annotations for this label (User-specific OR Public)
    saved_annotations = []
    from sqlalchemy import or_
    
    query = Annotation.query.filter(Annotation.set_id == set_id)
    
    if current_user.is_authenticated:
        query = query.filter(or_(Annotation.user_id == current_user.id, Annotation.is_public == True))
    else:
        query = query.filter(Annotation.is_public == True)
        
    user_annotations = query.all()
    
    for ann in user_annotations:
        saved_annotations.append({
            'id': str(ann.id), # Frontend expects string ID
            'section_number': ann.section_number,
            'question': ann.question,
            'answer': ann.answer,
            'keywords': json.loads(ann.keywords) if ann.keywords else [],
            'is_public': getattr(ann, 'is_public', False)
        })

    # Fetch metadata to get details for the title and FAERS search
    metadata = get_label_metadata(set_id, import_id=import_id)
    
    brand_name = metadata.get('brand_name') if metadata and metadata.get('brand_name') != 'N/A' else None
    generic_name = metadata.get('generic_name') if metadata and metadata.get('generic_name') != 'N/A' else None
    manufacturer_name = metadata.get('manufacturer_name') if metadata and metadata.get('manufacturer_name') != 'N/A' else None
    effective_time = metadata.get('effective_time') if metadata else ''
    label_format = metadata.get('label_format') if metadata else None
    ndc = metadata.get('ndc') if metadata else None
    application_number = metadata.get('application_number') if metadata else None
    version_number = metadata.get('version_number') if metadata else None
    document_type = metadata.get('document_type') if metadata else None
    has_boxed_warning = metadata.get('has_boxed_warning') if metadata else False
    
    clean_app_num = None
    if application_number and application_number != 'N/A':
        # Take the first one if multiple
        first_app = application_number.split(',')[0].strip()
        # Extract just the digits (e.g. ANDA077844 -> 077844)
        clean_app_num = ''.join(filter(str.isdigit, first_app))

    # Construct the display title: Brand Name - Generic Name, Company
    if brand_name:
        display_parts = [brand_name]
        if generic_name:
            display_parts.append(f"- {generic_name}")
        
        display_drug_name = " ".join(display_parts)
        if manufacturer_name:
            display_drug_name += f", {manufacturer_name}"
    else:
        # Fallback to query param or document title if metadata fetch fails
        display_drug_name = drug_name_from_query if drug_name_from_query else doc_title

    # Clean up name for FAERS search (take first brand name if multiple)
    faers_drug_name = brand_name if brand_name else display_drug_name
    if faers_drug_name and ',' in faers_drug_name:
        faers_drug_name = faers_drug_name.split(',')[0].strip()

    # Clean the original XML title (remove Highlights disclaimer)
    original_title = doc_title
    if "These highlights do not include all the information" in doc_title:
        original_title = re.sub(r'^These highlights do not include.*?safely and effectively\.\s*(See full prescribing information for.*?\.)?\s*', '', doc_title, flags=re.IGNORECASE | re.DOTALL).strip()

    if request.headers.get('Accept') == 'application/json' or request.args.get('json') == '1':
        return jsonify({
            'drug_name': display_drug_name,
            'brand_name': brand_name,
            'generic_name': generic_name,
            'original_title': original_title,
            'faers_drug_name': faers_drug_name,
            'manufacturer_name': manufacturer_name or '',
            'effective_time': effective_time,
            'label_format': label_format,
            'ndc': ndc,
            'application_number': application_number,
            'version_number': version_number,
            'document_type': document_type,
            'has_boxed_warning': has_boxed_warning,
            'clean_app_num': clean_app_num,
            'original_search': drug_name_from_query,
            'sections': sections, 
            'fallback_html': fallback_html, 
            'highlights': highlights, 
            'table_of_contents': table_of_contents,
            'label_xml_raw': label_xml_raw,
            'set_id': set_id,
            'metadata': metadata,
            'saved_annotations': saved_annotations,
            'user_id': current_user.id if current_user.is_authenticated else None
        })

    return render_template('results.html', 
                           drug_name=display_drug_name, 
                           brand_name=brand_name,
                           generic_name=generic_name,
                           original_title=original_title,
                           faers_drug_name=faers_drug_name,
                           manufacturer_name=manufacturer_name or '',
                           effective_time=effective_time,
                           label_format=label_format,
                           ndc=ndc,
                           application_number=application_number,
                           version_number=version_number,
                           document_type=document_type,
                           has_boxed_warning=has_boxed_warning,
                           clean_app_num=clean_app_num,
                           original_search=drug_name_from_query,
                           sections=sections, 
                           fallback_html=fallback_html, 
                           highlights=highlights, 
                           table_of_contents=table_of_contents,
                           label_xml_raw=label_xml_raw,
                           set_id=set_id,
                           metadata=metadata,
                           saved_annotations=saved_annotations)

@main_bp.route('/compare', methods=['GET', 'POST'])
def compare():
    """ Handles the comparison of multiple drug labels with deep, sorted comparison. """
    if request.method == 'POST':
        set_ids = request.form.getlist('set_ids')
        drug_name = request.form.get('drug_name')
        page = request.form.get('page', type=int)
        view = request.form.get('view')
        import_id = request.form.get('import_id')
    else: # GET
        set_ids = request.args.getlist('set_ids')
        drug_name = request.args.get('drug_name')
        page = request.args.get('page', type=int)
        view = request.args.get('view')
        import_id = request.args.get('import_id')

    if not set_ids:
        return jsonify({'error': 'Please select at least one label.'}), 400

    # Limit to 3 labels
    if len(set_ids) > 3:
        set_ids = set_ids[:3]

    # Perform format check for all selected labels before proceeding
    selected_labels_metadata = []
    formats = set()
    for set_id in set_ids:
        # Fetch full metadata for format check and summary panel
        metadata = get_label_metadata(set_id, import_id=import_id)
        if not metadata or not metadata.get('label_format'):
            return jsonify({
                'error': f'Label with SET ID {set_id} has an unsupported format and cannot be compared.',
                'unsupported_set_id': set_id
            }), 400
        
        selected_labels_metadata.append(metadata)
        formats.add(metadata.get('label_format'))

    error_msg = None
    if len(formats) > 1:
        # Keep only the format of the first label (primary)
        primary_format = selected_labels_metadata[0].get('label_format')
        valid_set_ids = []
        valid_metadata = []
        
        for meta in selected_labels_metadata:
            if meta.get('label_format') == primary_format:
                valid_set_ids.append(meta['set_id'])
                valid_metadata.append(meta)
        
        # Update the lists to continue processing with valid data
        set_ids = valid_set_ids
        selected_labels_metadata = valid_metadata
        
        error_msg = 'Mixing different label formats (PLR and non-PLR) is not supported. The incompatible label was not added.'

    labels_data = []
    all_section_keys = {} # Use a map to store unique keys (numeric ID for PLR, Title for non-PLR)
    comparison_format = selected_labels_metadata[0].get('label_format')

    # First, fetch, parse, and flatten all label data
    for set_id in set_ids:
        label_xml_raw = get_label_xml(set_id)
        if label_xml_raw:
            doc_title, sections, _, _, _ = parse_spl_xml(label_xml_raw)
            
            flat_sections = flatten_sections(sections)
            
            # Create a dictionary for faster section lookup by key
            sections_by_key = {}
            for s in flat_sections:
                if s.get('title'):
                    raw_title = s['title']
                    norm_title = normalize_title_text(raw_title)
                    
                    if comparison_format == 'PLR':
                        key = extract_numeric_section_id(raw_title)
                        if key:
                            # Aggregate content from children without their own numeric IDs
                            sections_by_key[key] = get_aggregate_content(s)
                            # Collect unique semantic titles for this numeric ID
                            if key not in all_section_keys:
                                all_section_keys[key] = {} # norm_title -> original_title
                            if norm_title not in all_section_keys[key]:
                                all_section_keys[key][norm_title] = raw_title
                    else: # non-PLR
                        key = norm_title # Use normalized title as key for non-PLR
                        if key:
                            sections_by_key[key] = s['content']
                            if key not in all_section_keys:
                                all_section_keys[key] = {}
                            if norm_title not in all_section_keys[key]:
                                all_section_keys[key][norm_title] = raw_title
            
            labels_data.append({
                'title': doc_title,
                'sections_by_key': sections_by_key
            })
            
    # Sort keys
    if comparison_format == 'PLR':
        sorted_keys = sorted(all_section_keys.keys(), key=lambda x: get_section_sort_key(x))
    else:
        sorted_keys = sorted(all_section_keys.keys()) # Alphabetical for non-PLR

    # Now, build the comparison data structure
    comparison_data = []
    htmldiffer = HtmlDiff() # Initialize HtmlDiff once, outside the loop

    def aggressive_normalize(lines):
        """Further normalizes text to ignore case and non-alphanumeric characters."""
        if not lines:
            return ""
        text = " ".join(lines).lower()
        return re.sub(r'[^a-z0-9]', '', text)

    for key in sorted_keys:
        contents = []
        for label in labels_data:
            contents.append(label['sections_by_key'].get(key))
        
        # Determine the display title
        # Pick the original titles associated with unique normalized versions
        display_titles = sorted(list(all_section_keys[key].values()))
        display_title = " / ".join(display_titles)

        # Semantically check if contents are the same
        normalized_contents = [tuple(normalize_text_for_diff(c)) if c else None for c in contents]
        
        # Aggressively normalize for the "Same" check to ignore tiny differences (case, punctuation)
        agg_normalized_contents = [aggressive_normalize(nc) for nc in normalized_contents]
        
        # Check if the section content is effectively empty across all labels
        is_empty = all(not nc for nc in agg_normalized_contents)

        # is_same is true only if:
        # 1. All labels have this section (no None)
        # 2. All aggressive normalized versions are identical
        has_all = all(nc is not None for nc in agg_normalized_contents)
        is_same = has_all and len(set(agg_normalized_contents)) == 1

        diff_html_output = None
        # Only generate diff if there are exactly two labels being compared
        # and they are not semantically identical according to the aggressive check
        if not is_same and len(contents) == 2:
            # If one is missing, we show it as a diff against an empty string
            plain_text1_lines = normalized_contents[0] if normalized_contents[0] is not None else []
            plain_text2_lines = normalized_contents[1] if normalized_contents[1] is not None else []
            
            diff_html_output = htmldiffer.make_table(
                plain_text1_lines, 
                plain_text2_lines, 
                fromdesc="1", 
                todesc="2",
                context=True # Show context lines around differences
            )

            if diff_html_output:
                # Clean up difflib output for better wrapping
                diff_html_output = diff_html_output.replace('nowrap="nowrap"', '')
                diff_html_output = diff_html_output.replace('&nbsp;', ' ')
                
                # Remove colgroup tags as they interfere with our custom styling
                diff_html_output = re.sub(r'<colgroup>.*?</colgroup>', '', diff_html_output)
                
                # Replace the default thead with "1 / Content / 2 / Content"
                new_thead = (
                    '<thead><tr>'
                    '<th class="diff_next"></th>'
                    '<th class="diff_header">1</th><th class="diff_header">Content</th>'
                    '<th class="diff_next"></th>'
                    '<th class="diff_header">2</th><th class="diff_header">Content</th>'
                    '</tr></thead>'
                )
                diff_html_output = re.sub(r'<thead>.*?</thead>', new_thead, diff_html_output, flags=re.DOTALL)

        comparison_data.append({
            'title': display_title,
            'key': key,
            'nesting_level': key.count('.') if comparison_format == 'PLR' else 0,
            'contents': contents, 
            'is_same': is_same,
            'is_empty': is_empty,
            'diff_html': diff_html_output
        })
    
    # Generate a default title for the comparison
    labels_titles = [m.get('brand_name') or m.get('set_id') for m in selected_labels_metadata]
    if len(labels_titles) > 1:
        comparison_title = " vs ".join(labels_titles)
    else:
        comparison_title = f"Label: {labels_titles[0]}"

    user_favorites = []
    if current_user.is_authenticated:
        user_favorites = Favorite.query.filter_by(user_id=current_user.id).order_by(Favorite.timestamp.desc()).all()

    return render_template('compare.html', 
                           labels=[ld['title'] for ld in labels_data], 
                           comparison_data=comparison_data,
                           selected_labels_metadata=selected_labels_metadata,
                           drug_name=drug_name,
                           page=page,
                           view=view,
                           current_set_ids=set_ids, # Pass set_ids to template
                           comparison_title=comparison_title, # Pass title
                           user_favorites=user_favorites, # Pass favorites
                           error=error_msg) # Pass error message

@main_bp.route('/my_labelings')
@login_required
def my_labelings():
    return render_template('my_labelings.html')

@main_bp.route('/my_favorites')
@login_required
def my_favorites():
    return redirect(url_for('main.my_labelings'))

@main_bp.route('/preferences', methods=['POST'])
@login_required
def preferences():
    current_user.ai_provider = request.form.get('ai_provider')
    current_user.custom_gemini_key = request.form.get('custom_gemini_key')
    current_user.openai_api_key = request.form.get('openai_api_key')
    current_user.openai_base_url = request.form.get('openai_base_url')
    current_user.openai_model_name = request.form.get('openai_model_name')
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Preferences saved successfully!'})

