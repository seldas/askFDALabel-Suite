from flask import Blueprint, request, jsonify, redirect, url_for
from flask_login import login_required, current_user
import json, re, os
import uuid
import requests
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
import logging
import hashlib

from database import (
    db, User, Project, Favorite, FavoriteComparison, Annotation, 
    LabelAnnotation, DiliAssessment, DictAssessment, DiriAssessment, ToxAgent, ComparisonSummary,
    MeddraPT, MeddraMDHIER, MeddraSOC, MeddraHLT, MeddraLLT,
    ProjectAeReport, ProjectAeReportDetail, AeAiAssessment
)
import threading
from dashboard.services.fda_client import get_label_metadata, get_label_xml, get_faers_data, find_labels, find_labels_by_set_ids, get_label_counts, get_rich_metadata_by_generic
from dashboard.services.ai_handler import chat_with_document, summarize_comparison, generate_assessment, get_search_helper_response
from dashboard.services.xml_handler import extract_metadata_from_xml
from dashboard.services.pgx_handler import run_pgx_assessment
from dashboard.prompts import (
    DILI_prompt, DICT_prompt, DIRI_prompt, 
    DILI_PT_TERMS, DICT_PT_TERMS, DIRI_PT_TERMS)
from dashboard.config import Config
from dashboard.services.fdalabel_db import FDALabelDBService
from dashboard.services.deep_dive_service import DeepDiveService
from sqlalchemy import func
from dashboard.services.meddra_matcher import scan_label_for_meddra

logger = logging.getLogger(__name__)
api_bp = Blueprint('api', __name__)

@api_bp.route('/deep_dive/peers_count/<set_id>')
def get_peers_count(set_id):
    """
    Fetches counts of peer labels (same name and same EPC) from various sources.
    Source can be 'local', 'oracle', or 'openfda'.
    If the current set_id is missing EPC data, it searches by Generic Name
    to borrow metadata from a richer peer record.
    """
    source = request.args.get('source', 'openfda').lower()
    
    # 1. Get basic metadata for the current set_id
    meta = get_label_metadata(set_id)
    if not meta:
        return jsonify({'error': 'Label metadata not found'}), 404
    
    generic_name = meta.get('generic_name')
    epc = meta.get('epc')
    
    # 2. BORROWING STRATEGY: If EPC is missing, search by Generic Name for a rich peer
    if not epc or epc.lower() == 'n/a':
        logger.info(f"EPC missing for {set_id}. Searching by generic name: {generic_name}")
        rich_meta = get_rich_metadata_by_generic(generic_name)
        if rich_meta:
            epc = rich_meta.get('epc')
            logger.info(f"Borrowed EPC for {generic_name}: {epc}")

    # Final cleanup
    if generic_name == 'n/a' or generic_name == 'Unknown Generic':
        generic_name = None
    if epc == 'n/a':
        epc = None

    results = {"source": source, "names": [], "epcs": []}

    # Helper to get detailed counts for each term
    def get_detailed_counts(name_str, epc_str, count_fn):
        names = [n.strip() for n in (name_str or "").split(',') if n.strip() and n.strip().lower() != 'n/a']
        epcs = [e.strip() for e in (epc_str or "").split(',') if e.strip() and e.strip().lower() != 'n/a']
        
        name_list = []
        for n in names:
            c = count_fn(generic_name=n)
            name_list.append({"term": n, "count": c.get('generic_count', 0)})
            
        epc_list = []
        for e in epcs:
            c = count_fn(epc=e)
            epc_list.append({"term": e, "count": c.get('epc_count', 0)})
            
        return name_list, epc_list

    try:
        if source == 'openfda':
            n_list, e_list = get_detailed_counts(generic_name, epc, get_label_counts)
        else:
            n_list, e_list = get_detailed_counts(generic_name, epc, FDALabelDBService.get_label_counts)
        
        results["names"] = n_list
        results["epcs"] = e_list
            
        return jsonify(results)
    except Exception as e:
        logger.error(f"Error in get_peers_count: {e}")
        return jsonify({'error': str(e)}), 500

@api_bp.route('/deep_dive/analysis/<set_id>')
def get_deep_dive_analysis(set_id):
    """
    Runs the Phase 2 statistical analysis (TF-IDF) for the label sections.
    """
    source = request.args.get('source', 'openfda').lower()
    generic_names = request.args.get('generic_names')
    epcs = request.args.get('epcs')
    
    try:
        analysis = DeepDiveService.get_comparison_analysis(
            target_set_id=set_id,
            source=source,
            generic_names=generic_names,
            epcs=epcs
        )
        return jsonify(analysis)
    except Exception as e:
        logger.error(f"Error in get_deep_dive_analysis: {e}")
        return jsonify({'error': str(e)}), 500

@api_bp.route('/meddra/scan_label/<set_id>')
def scan_label_meddra(set_id):
    # 1. Get XML content
    xml_content = get_label_xml(set_id)
    if not xml_content:
        return jsonify({'error': 'Label not found'}), 404

    # 2. Extract plain text (simple stripping for matching purposes)
    try:
        root = ET.fromstring(xml_content.encode('ascii', 'ignore').decode('ascii'))
        text_content = "".join(root.itertext())
    except Exception as e:
        logger.error(f"XML Parse error for MedDRA scan: {e}")
        return jsonify({'error': 'Failed to parse label'}), 500

    # 3. Scan
    found_terms = scan_label_for_meddra(text_content)
    
    # 4. Enrich with SOC info
    enriched_terms = []
    if found_terms:
        # Create a mock list for enrichment function
        mock_list = [{'term': t} for t in found_terms]
        enriched_list = enrich_faers_with_meddra(mock_list)
        
        # Transform back to simpler structure if needed, or just return
        enriched_terms = enriched_list

    return jsonify({'found_terms': enriched_terms, 'count': len(enriched_terms)})

@api_bp.route('/meddra/profile/<set_id>')
def get_label_meddra_profile(set_id):
    """
    Returns a JSON profile of all MedDRA terms found in specific SPL sections.
    This endpoint is public (no login required).
    """
    xml_content = get_label_xml(set_id)
    if not xml_content:
        return jsonify({'error': 'Label not found'}), 404

    target_sections = {
        '34066-1': 'Boxed Warning',
        '34070-3': 'Contraindications',
        '34071-1': 'Warnings and Precautions',
        '43685-7': 'Warnings and Precautions',
        '34084-4': 'Adverse Reactions'
    }

    profile_data = {
        'metadata': {
            'set_id': set_id,
            'generated_at': datetime.utcnow().isoformat()
        },
        'sections': [],
        'all_found_terms': []
    }

    try:
        ns = {'v3': 'urn:hl7-org:v3'}
        # Clean XML string from any non-ascii characters that might cause ET issues
        clean_xml = xml_content.encode('ascii', 'ignore').decode('ascii')
        root = ET.fromstring(clean_xml)
        
        all_terms_set = set()
        
        for section in root.findall(".//v3:section", ns):
            code_el = section.find("v3:code", ns)
            if code_el is not None:
                code_val = code_el.get('code')
                if code_val in target_sections:
                    section_name = target_sections[code_val]
                    # Extract text content from this section only
                    # We iterate through children to avoid getting text from nested subsections twice
                    # but itertext is simpler for initial extraction
                    text_content = "".join(section.itertext()).strip()
                    
                    if text_content:
                        found = scan_label_for_meddra(text_content)
                        if found:
                            # Enrich found terms for this section
                            mock_list = [{'term': t} for t in found]
                            enriched = enrich_faers_with_meddra(mock_list)
                            
                            profile_data['sections'].append({
                                'section_code': code_val,
                                'section_name': section_name,
                                'terms': enriched
                            })
                            
                            for t in found:
                                all_terms_set.add(t)

        # Final summary enrichment
        if all_terms_set:
            summary_list = [{'term': t} for t in all_terms_set]
            profile_data['all_found_terms'] = enrich_faers_with_meddra(summary_list)

        # Add label metadata to the final response
        meta = extract_metadata_from_xml(xml_content)
        if meta:
            profile_data['metadata'].update({
                'brand_name': meta.get('brand_name'),
                'generic_name': meta.get('generic_name'),
                'manufacturer': meta.get('manufacturer_name'),
                'effective_time': meta.get('effective_time')
            })

        return jsonify(profile_data)

    except Exception as e:
        logger.error(f"Error generating MedDRA profile for {set_id}: {e}")
        return jsonify({'error': str(e)}), 500

def enrich_faers_with_meddra(faers_list):
    """
    Enriches a list of FAERS results (dict with 'term') with SOC and HLT info.
    Handles both PTs and LLTs.
    """
    if not faers_list:
        return faers_list

    # Extract terms (unique, upper case)
    all_terms = set(item.get('term').upper() for item in faers_list if item.get('term'))
    if not all_terms:
        return faers_list

    enrichment_map = {}
    
    # 1. Try to match as PTs
    pt_results = db.session.query(
        MeddraPT.pt_name,
        MeddraMDHIER.soc_name,
        MeddraMDHIER.hlt_name,
        MeddraMDHIER.soc_abbrev
    ).join(
        MeddraMDHIER, MeddraPT.pt_code == MeddraMDHIER.pt_code
    ).filter(
        MeddraMDHIER.primary_soc_fg == 'Y', 
        func.upper(MeddraPT.pt_name).in_(all_terms)
    ).all()

    found_pts = set()
    for pt_name, soc_name, hlt_name, soc_abbrev in pt_results:
        u_name = pt_name.upper()
        enrichment_map[u_name] = {
            'soc': soc_name, 
            'hlt': hlt_name,
            'soc_abbrev': soc_abbrev
        }
        found_pts.add(u_name)

    # 2. Identify missing terms (potential LLTs)
    missing_terms = all_terms - found_pts
    
    if missing_terms:
        llt_results = db.session.query(
            MeddraLLT.llt_name,
            MeddraMDHIER.soc_name,
            MeddraMDHIER.hlt_name,
            MeddraMDHIER.soc_abbrev
        ).join(
            MeddraMDHIER, MeddraLLT.pt_code == MeddraMDHIER.pt_code
        ).filter(
            MeddraMDHIER.primary_soc_fg == 'Y',
            func.upper(MeddraLLT.llt_name).in_(missing_terms)
        ).all()

        for llt_name, soc_name, hlt_name, soc_abbrev in llt_results:
            u_name = llt_name.upper()
            enrichment_map[u_name] = {
                'soc': soc_name, 
                'hlt': hlt_name,
                'soc_abbrev': soc_abbrev
            }

    # 3. Update original list
    for item in faers_list:
        term = item.get('term', '').upper()
        if term in enrichment_map:
            item['soc'] = enrichment_map[term]['soc']
            item['hlt'] = enrichment_map[term]['hlt']
            item['soc_abbrev'] = enrichment_map[term]['soc_abbrev']
        else:
            item['soc'] = 'N/A'
            item['hlt'] = 'N/A'
            item['soc_abbrev'] = ''
            
    return faers_list

# --- AI Chat ---
@api_bp.route('/ai_chat', methods=['POST'])
def ai_chat():
    data = request.get_json()
    user_message = data.get('message')
    history = data.get('history', [])
    xml_content = data.get('xml_content')
    chat_type = data.get('chat_type', 'general')

    if not user_message or not xml_content:
        return jsonify({'error': 'Missing message or xml_content'}), 400

    try:
        user_obj = current_user._get_current_object() if current_user.is_authenticated else None
        response_text = chat_with_document(user_obj, user_message, history, xml_content, chat_type)
        return jsonify({'response': response_text})
    except Exception as e:
        logger.error(f"Error calling AI API: {e}, {current_user}, {chat_type}")
        return jsonify({'error': f'AI API error: {str(e)}'}), 500

@api_bp.route('/ai_search_help', methods=['POST'])
def ai_search_help():
    data = request.get_json()
    user_message = data.get('message')
    history = data.get('history', [])

    if not user_message:
        return jsonify({'error': 'Missing message'}), 400

    try:
        user_obj = current_user._get_current_object() if current_user.is_authenticated else None
        response_json = get_search_helper_response(user_obj, user_message, history)
        # Verify it's valid JSON
        try:
            # It might come back with markdown code blocks ```json ... ```
            cleaned = response_json.replace('```json', '').replace('```', '').strip()
            parsed = json.loads(cleaned)
            return jsonify(parsed)
        except:
            if '```json' in response_json:
                cleaned = re.search('''```json(.*?)```''', response_json, re.DOTALL).group(1).strip()
                parsed = json.loads(cleaned)
                return jsonify(parsed)
            else:
                # Fallback if AI didn't return valid JSON
                return jsonify({
                    "reply": response_json,
                    "suggested_term": None,
                    "is_final": False
                })
    except Exception as e:
        logger.error(f"Error calling AI Search Helper: {e}")
        return jsonify({'error': f'AI API error: {str(e)}'}), 500

@api_bp.route('/search_count', methods=['GET'])
def search_count():
    query = request.args.get('q')
    if not query:
        return jsonify({'count': 0, 'source': 'Unknown'})
    
    try:
        # Use find_labels with limit=1 just to get the total count
        _, total = find_labels(query, skip=0, limit=1)
        
        # Determine source
        from dashboard.services.fdalabel_db import FDALabelDBService
        if FDALabelDBService.is_internal():
            source = "FDALabel"
        elif FDALabelDBService.is_local():
            source = "LocalDB"
        else:
            source = "OpenFDA"
        
        return jsonify({'count': total, 'source': source})
    except Exception as e:
        logger.error(f"Error fetching search count: {e}")
        return jsonify({'count': 0, 'source': 'Error'})



@api_bp.route('/ai_compare_summary', methods=['POST'])
def ai_compare_summary():
    data = request.get_json()
    set_ids = data.get('set_ids', [])
    force_refresh = data.get('force_refresh', False)
    generate_if_missing = data.get('generate_if_missing', True)
    
    ids_hash = None
    ids_str = None

    # 1. Check Cache
    if set_ids:
        sorted_ids = sorted(set_ids)
        ids_str = json.dumps(sorted_ids)
        ids_hash = hashlib.sha256(ids_str.encode('utf-8')).hexdigest()
        
        if not force_refresh:
            cached = ComparisonSummary.query.filter_by(set_ids_hash=ids_hash).first()
            if cached:
                return jsonify({'summary': cached.summary_content, 'cached': True})

    if not generate_if_missing:
        return jsonify({'summary': None, 'cached': False})

    # 2. Generate
    differing_sections_data = data.get('differing_sections', [])
    label1_name = data.get('label1_name', 'Label A')
    label2_name = data.get('label2_name', 'Label B')

    if not differing_sections_data:
        return jsonify({'error': 'No differing sections provided for AI summary.'}), 400

    try:
        user_obj = current_user._get_current_object() if current_user.is_authenticated else None
        response_text = summarize_comparison(user_obj, differing_sections_data, label1_name, label2_name)
        
        # 3. Save to Cache
        if ids_hash:
            # Upsert
            existing = ComparisonSummary.query.filter_by(set_ids_hash=ids_hash).first()
            if existing:
                existing.summary_content = response_text
                existing.timestamp = datetime.utcnow()
            else:
                new_summary = ComparisonSummary(
                    set_ids_hash=ids_hash,
                    set_ids=ids_str,
                    summary_content=response_text
                )
                db.session.add(new_summary)
            db.session.commit()

        return jsonify({'summary': response_text, 'cached': False})
    except Exception as e:
        logger.error(f"Error calling AI API for comparison summary: {e}")
        return jsonify({'error': f'AI API error (Comparison): {str(e)}'}), 500


# --- Annotations ---
@api_bp.route('/save_annotation', methods=['POST'])
@login_required
def save_annotation():
    data = request.get_json()
    set_id = data.get('set_id')
    section_num = data.get('section_number', 'TOP') 
    question = data.get('question')
    answer = data.get('answer')
    keywords = data.get('keywords', [])
    is_public = data.get('is_public', False)
    
    if not set_id:
        return jsonify({'error': 'Missing set_id'}), 400
    
    if not question or not answer:
        return jsonify({'error': 'Missing question or answer'}), 400
        
    new_annotation = Annotation(
        user_id=current_user.id,
        set_id=set_id,
        section_number=str(section_num) if section_num else 'TOP',
        question=question,
        answer=answer,
        keywords=json.dumps(keywords),
        is_public=is_public
    )
    
    db.session.add(new_annotation)
    db.session.commit()
    
    return jsonify({'success': True, 'id': str(new_annotation.id)})

@api_bp.route('/delete_annotation', methods=['POST'])
@login_required
def delete_annotation():
    data = request.get_json()
    annotation_id = data.get('id')
    
    if not annotation_id:
        return jsonify({'error': 'Missing id'}), 400
        
    annotation = Annotation.query.filter_by(id=int(annotation_id), user_id=current_user.id).first()
    
    if annotation:
        db.session.delete(annotation)
        db.session.commit()
        return jsonify({'success': True})
        
    return jsonify({'error': 'Annotation not found or permission denied'}), 404

# --- Favorites ---
@api_bp.route('/toggle_favorite', methods=['POST'])
@login_required
def toggle_favorite():
    data = request.get_json()
    set_id = data.get('set_id')
    import_id = data.get('import_id')

    if not set_id:
        return jsonify({'error': 'Missing set_id'}), 400

    # LITERALLY always use the 'Favorite' project for toggling
    target_project = Project.query.filter_by(owner_id=current_user.id, title="Favorite").first()
    if not target_project:
        target_project = Project(title="Favorite", owner_id=current_user.id, display_order=0)
        db.session.add(target_project)
        db.session.commit()

    project_id = target_project.id

    # Check PROJECT-WIDE
    favorite = Favorite.query.filter_by(set_id=set_id, project_id=project_id, user_id=current_user.id).first()
    
    if favorite:
        # If it already exists, remove it (toggle behavior)
        db.session.delete(favorite)
        db.session.commit()
        return jsonify({'success': True, 'is_favorite': False})
    else:
        # Before adding, ensure we don't have a race condition or existing duplicate
        # (Though filter_by above should catch it, we stay strict)
        existing = Favorite.query.filter_by(set_id=set_id, project_id=project_id, user_id=current_user.id).first()
        if existing:
             return jsonify({'success': True, 'is_favorite': True})
            
        meta = get_label_metadata(set_id, import_id=import_id)
        if not meta:
            return jsonify({'error': 'Could not fetch label metadata'}), 404

        # REFINEMENT: If brand or manufacturer is 'n/a' or 'N/A', try extracting from XML
        brand = meta.get('brand_name', 'n/a')
        manufacturer = meta.get('manufacturer_name', 'n/a')
        
        if brand.lower() in ['n/a', 'unknown drug'] or manufacturer.lower() in ['n/a', 'unknown manufacturer']:
            from dashboard.services.fda_client import get_label_xml
            from dashboard.services.xml_handler import extract_metadata_from_xml
            xml_raw = get_label_xml(set_id)
            if xml_raw:
                xml_meta = extract_metadata_from_xml(xml_raw)
                if xml_meta:
                    if brand.lower() in ['n/a', 'unknown drug'] and xml_meta.get('brand_name') and xml_meta.get('brand_name') != 'Unknown Drug':
                        brand = xml_meta['brand_name']
                    if manufacturer.lower() in ['n/a', 'unknown manufacturer'] and xml_meta.get('manufacturer_name') and xml_meta.get('manufacturer_name') != 'Unknown Manufacturer':
                        manufacturer = xml_meta['manufacturer_name']

        new_favorite = Favorite(
            user_id=current_user.id, 
            project_id=project_id,
            set_id=set_id, 
            brand_name=brand,
            generic_name=meta.get('generic_name', 'n/a'),
            manufacturer_name=manufacturer,
            market_category=meta.get('market_category', 'n/a'),
            application_number=meta.get('application_number', 'n/a'),
            ndc=meta.get('ndc', 'n/a'),
            effective_time=meta.get('effective_time', 'n/a'),
            # New columns
            active_ingredients=meta.get('active_ingredients', 'n/a'),
            labeling_type=meta.get('labeling_type', 'n/a'),
            dosage_forms=meta.get('dosage_forms', 'n/a'),
            routes=meta.get('routes', 'n/a'),
            epc=meta.get('epc', 'n/a'),
            fdalabel_link=f"https://nctr-crs.fda.gov/fdalabel/ui/search/spl/{set_id}",
            dailymed_spl_link=f"https://dailymed.nlm.nih.gov/dailymed/lookup.cfm?setid={set_id}",
            dailymed_pdf_link=f"https://dailymed.nlm.nih.gov/dailymed/getpdf.cfm?setid={set_id}",
            product_type=meta.get('product_type', 'n/a'),
            label_format=meta.get('label_format', 'n/a'),
            source=meta.get('source', 'n/a')
        )
        db.session.add(new_favorite)
        db.session.commit()
        return jsonify({'success': True, 'is_favorite': True})

@api_bp.route('/check_favorite/<set_id>')
def check_favorite(set_id):
    if not current_user.is_authenticated:
        return jsonify({'is_favorite': False})
    
    # Yellow star depends ONLY on the 'Favorite' project
    target_project = Project.query.filter_by(owner_id=current_user.id, title="Favorite").first()
    if not target_project:
        return jsonify({'is_favorite': False})

    favorite = Favorite.query.filter_by(set_id=set_id, project_id=target_project.id).first()
    return jsonify({'is_favorite': bool(favorite)})

@api_bp.route('/toggle_favorite_comparison', methods=['POST'])
@login_required
def toggle_favorite_comparison():
    data = request.get_json()
    set_ids = data.get('set_ids') # list
    title = data.get('title')
    project_id = data.get('project_id')

    if not set_ids:
        return jsonify({'error': 'Missing set_ids'}), 400

    set_ids.sort()
    set_ids_json = json.dumps(set_ids)

    if not project_id:
        default_proj = Project.query.filter_by(owner_id=current_user.id, title="Favorite").first()
        if default_proj:
            project_id = default_proj.id

    # Check PROJECT-WIDE
    favorite = FavoriteComparison.query.filter_by(set_ids=set_ids_json, project_id=project_id).first()
    
    # Check permissions
    project = Project.query.get(project_id)
    if not project or (project.owner_id != current_user.id and current_user not in project.members):
         return jsonify({'error': 'Unauthorized'}), 403

    if favorite:
        db.session.delete(favorite)
        db.session.commit()
        return jsonify({'success': True, 'is_favorite': False})
    else:
        new_favorite = FavoriteComparison(user_id=current_user.id, set_ids=set_ids_json, title=title, project_id=project_id)
        db.session.add(new_favorite)
        db.session.commit()
        return jsonify({'success': True, 'is_favorite': True})

@api_bp.route('/check_favorite_comparison', methods=['POST'])
def check_favorite_comparison():
    if not current_user.is_authenticated:
        return jsonify({'is_favorite': False})
    
    data = request.get_json()
    set_ids = data.get('set_ids')
    project_id = data.get('project_id')

    if not set_ids:
        return jsonify({'is_favorite': False})

    set_ids.sort()
    set_ids_json = json.dumps(set_ids)
    
    if not project_id:
        default_proj = Project.query.filter_by(owner_id=current_user.id, title="Favorite").first()
        if default_proj:
            project_id = default_proj.id

    # Check PROJECT-WIDE
    favorite = FavoriteComparison.query.filter_by(set_ids=set_ids_json, project_id=project_id).first()
    return jsonify({'is_favorite': bool(favorite)})

# --- Label Annotations ---
from flask import request, jsonify
from sqlalchemy import or_

# --- Label Annotations ---
@api_bp.route('/annotations/save', methods=['POST'])
@login_required
def save_label_annotation():
    data = request.get_json(silent=True) or {}

    raw_project_id = data.get('project_id', None)
    project_id = None
    if raw_project_id not in (None, '', 'null'):
        try:
            project_id = int(raw_project_id)
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid project_id'}), 400

    set_id = data.get('set_id')
    section_id = data.get('section_id')
    start_offset = data.get('start_offset')
    end_offset = data.get('end_offset')
    selected_text = data.get('selected_text')
    annotation_type = data.get('annotation_type')
    color = data.get('color')
    comment = data.get('comment')

    required_fields = {
        'set_id': set_id,
        'section_id': section_id,
        'start_offset': start_offset,
        'end_offset': end_offset,
        'selected_text': selected_text,
        'annotation_type': annotation_type
    }
    missing_fields = [
        k for k, v in required_fields.items()
        if v is None or (isinstance(v, str) and v.strip() == '')
    ]
    if missing_fields:
        return jsonify({'error': f"Missing required fields: {', '.join(missing_fields)}"}), 400

    try:
        start_offset = int(start_offset)
        end_offset = int(end_offset)
    except (TypeError, ValueError):
        return jsonify({'error': 'start_offset and end_offset must be integers'}), 400

    if start_offset < 0 or end_offset <= start_offset:
        return jsonify({'error': 'Invalid offset range'}), 400

    # Validate project access only if project_id provided
    if project_id is not None:
        project = Project.query.get(project_id)
        if not project:
            return jsonify({'error': 'Project not found'}), 404
        if project.owner_id != current_user.id and current_user not in project.members:
            return jsonify({'error': 'Unauthorized access to project'}), 403

    new_ann = LabelAnnotation(
        project_id=project_id,  # None => global
        set_id=set_id,
        user_id=current_user.id,
        section_id=section_id,
        start_offset=start_offset,
        end_offset=end_offset,
        selected_text=selected_text,
        annotation_type=annotation_type,
        color=color,
        comment=comment
    )
    db.session.add(new_ann)
    db.session.commit()

    return jsonify({
        'success': True,
        'id': new_ann.id,
        'project_id': new_ann.project_id,
        'is_global': new_ann.project_id is None,
        'username': current_user.username,
        'created_at': new_ann.created_at.isoformat()
    })

from sqlalchemy import or_

@api_bp.route('/annotations/get/<set_id>')
@login_required
def get_label_annotations(set_id):
    project_id = request.args.get('project_id', type=int)

    # If a project_id is provided, validate access to that project
    if project_id:
        project = Project.query.get(project_id)
        if not project:
            return jsonify({'error': 'Project not found'}), 404

        if project.owner_id != current_user.id and current_user not in project.members:
            return jsonify({'error': 'Unauthorized access to project'}), 403

        # Return BOTH:
        # - project-specific annotations (project_id == project_id)
        # - global annotations for this user (project_id IS NULL)
        annotations = (
            LabelAnnotation.query
            .filter(LabelAnnotation.set_id == set_id)
            .filter(LabelAnnotation.user_id == current_user.id)
            .filter(or_(LabelAnnotation.project_id == project_id,
                        LabelAnnotation.project_id.is_(None)))
            .all()
        )
    else:
        # No project specified: return only global annotations for this user
        annotations = (
            LabelAnnotation.query
            .filter(LabelAnnotation.set_id == set_id)
            .filter(LabelAnnotation.user_id == current_user.id)
            .filter(LabelAnnotation.project_id.is_(None))
            .all()
        )

    return jsonify({
        'annotations': [{
            'id': ann.id,
            'section_id': ann.section_id,
            'start_offset': ann.start_offset,
            'end_offset': ann.end_offset,
            'selected_text': ann.selected_text,
            'annotation_type': ann.annotation_type,
            'color': ann.color,
            'comment': ann.comment,
            'user_id': ann.user_id,
            'username': ann.user.username if ann.user else None,
            'created_at': ann.created_at.isoformat() if ann.created_at else None,
            'project_id': ann.project_id,  # helpful for debugging/UI
            'is_global': ann.project_id is None
        } for ann in annotations]
    })


@api_bp.route('/annotations/delete', methods=['POST'])
@login_required
def delete_label_annotation():
    data = request.json
    ann_id = data.get('id')
    
    annotation = LabelAnnotation.query.get(ann_id)
    if not annotation:
        return jsonify({'error': 'Annotation not found'}), 404
        
    if annotation.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized. Only the creator can delete this annotation.'}), 403
        
    db.session.delete(annotation)
    db.session.commit()
    
    return jsonify({'success': True})

# --- Projects ---
@api_bp.route('/projects', methods=['GET', 'POST'])
@login_required
def api_projects():
    if request.method == 'POST':
        data = request.json
        title = data.get('title')
        if not title:
            return jsonify({'error': 'Title is required'}), 400
        
        count = Project.query.filter_by(owner_id=current_user.id).count()
        if count >= 100:
            return jsonify({'error': 'Project limit reached (Max 100).'}), 403
            
        max_order = db.session.query(db.func.max(Project.display_order)).filter_by(owner_id=current_user.id).scalar() or 0
        
        new_proj = Project(title=title, description=data.get('description'), owner_id=current_user.id, display_order=max_order + 1)
        db.session.add(new_proj)
        db.session.commit()
        return jsonify({'success': True, 'project': {'id': new_proj.id, 'title': new_proj.title}})
    
    not_grouped = Project.query.filter_by(owner_id=current_user.id, title="Favorite").first()
    
    other_owned = Project.query.filter(
        Project.owner_id == current_user.id, 
        Project.title != "Favorite"
    ).order_by(Project.display_order.asc(), Project.created_at.asc()).all()
    
    shared = current_user.shared_projects
    
    projects = []
    if not_grouped:
        projects.append({
            'id': not_grouped.id, 
            'title': not_grouped.title, 
            'role': 'owner', 
            'count': len(not_grouped.favorites) + len(not_grouped.comparisons), 
            'is_default': True,
            'is_mutable': False # Favorite project is a permanent workspace and cannot be removed
        })
        
    for p in other_owned:
        projects.append({
            'id': p.id, 
            'title': p.title, 
            'role': 'owner', 
            'count': len(p.favorites) + len(p.comparisons), 
            'is_default': False,
            'is_mutable': True # Non-favorite projects can be removed freely by the owner
        })
        
    for p in shared:
        projects.append({
            'id': p.id, 
            'title': p.title, 
            'role': 'contributor', 
            'count': len(p.favorites) + len(p.comparisons), 
            'is_default': False,
            'is_mutable': False
        })
        
    return jsonify({'projects': projects})

@api_bp.route('/projects/<int:project_id>', methods=['PUT', 'DELETE'])
@login_required
def api_project_detail(project_id):
    project = Project.query.get_or_404(project_id)
    
    if project.owner_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403

    if request.method == 'DELETE':
        # Protect the 'Favorite' project from deletion
        if project.title == "Favorite":
            return jsonify({'error': 'The Favorite project is a permanent workspace and cannot be removed.'}), 403
        
        db.session.delete(project)
        db.session.commit()
        return jsonify({'success': True})
        
    if request.method == 'PUT':
        # Protect the 'Favorite' project from being renamed
        if project.title == "Favorite":
             return jsonify({'error': 'The Favorite project cannot be renamed.'}), 403
             
        data = request.json
        if 'title' in data:
            project.title = data['title']
        if 'description' in data:
            project.description = data['description']
        db.session.commit()
        return jsonify({'success': True})

@api_bp.route('/favorites_data')
@login_required
def api_my_favorites():
    project_id = request.args.get('project_id', type=int)
    user_obj = current_user._get_current_object()
    
    if project_id:
        project = Project.query.get(project_id)
        if not project or (project.owner_id != user_obj.id and user_obj not in project.members):
             return jsonify({'error': 'Unauthorized'}), 403
             
        # DEDUPLICATION LOGIC:
        # If this is the owner accessing, we clean up any duplicate set_ids that might have leaked in
        if project.owner_id == user_obj.id:
            all_favs = Favorite.query.filter_by(project_id=project_id).all()
            seen_set_ids = set()
            duplicates_removed = False
            for f in all_favs:
                if f.set_id in seen_set_ids:
                    db.session.delete(f)
                    duplicates_removed = True
                else:
                    seen_set_ids.add(f.set_id)
            if duplicates_removed:
                db.session.commit()

        favorites = Favorite.query.filter_by(project_id=project_id).order_by(Favorite.timestamp.desc()).all()
        favorite_comparisons = FavoriteComparison.query.filter_by(project_id=project_id).order_by(FavoriteComparison.timestamp.desc()).all()
        
        return jsonify({
            'favorites': [{
                'set_id': fav.set_id,
                'brand_name': fav.brand_name,
                'generic_name': fav.generic_name,
                'manufacturer_name': fav.manufacturer_name,
                'market_category': fav.market_category,
                'application_number': fav.application_number,
                'ndc': fav.ndc,
                'effective_time': fav.effective_time,
                'timestamp': fav.timestamp.isoformat(),
                'added_by': fav.user.username
            } for fav in favorites],
            'comparisons': [{
                'id': comp.id,
                'set_ids': json.loads(comp.set_ids),
                'title': comp.title,
                'timestamp': comp.timestamp.isoformat(),
                'added_by': comp.user.username
            } for comp in favorite_comparisons],
            'duplicates_removed': duplicates_removed
        })
    else:
        # Global view (all projects) - less common but same logic
        favorites = Favorite.query.filter_by(user_id=user_obj.id).order_by(Favorite.timestamp.desc()).all()
        favorite_comparisons = FavoriteComparison.query.filter_by(user_id=user_obj.id).order_by(FavoriteComparison.timestamp.desc()).all()
        
        return jsonify({
            'favorites': [{
                'set_id': fav.set_id,
                'brand_name': fav.brand_name,
                'generic_name': fav.generic_name,
                'manufacturer_name': fav.manufacturer_name,
                'market_category': fav.market_category,
                'application_number': fav.application_number,
                'ndc': fav.ndc,
                'effective_time': fav.effective_time,
                'timestamp': fav.timestamp.isoformat(),
                'added_by': fav.user.username
            } for fav in favorites],
            'comparisons': [{
                'id': comp.id,
                'set_ids': json.loads(comp.set_ids),
                'title': comp.title,
                'timestamp': comp.timestamp.isoformat(),
                'added_by': comp.user.username
            } for comp in favorite_comparisons],
            'duplicates_removed': False
        })

@api_bp.route('/check_favorites_batch', methods=['POST'])
def api_check_favorites_batch():
    if not current_user.is_authenticated:
        return jsonify({})
    
    data = request.get_json()
    set_ids = data.get('set_ids', [])
    project_id = data.get('project_id')
    
    if not set_ids:
        return jsonify({})

    if not project_id:
        default_proj = Project.query.filter_by(owner_id=current_user.id, title="Favorite").first()
        if default_proj:
            project_id = default_proj.id
    
    # Check PROJECT-WIDE
    favorites = Favorite.query.filter(
        Favorite.project_id == project_id,
        Favorite.set_id.in_(set_ids)
    ).all()
    
    found_ids = {fav.set_id: True for fav in favorites}
    
    result = {}
    for sid in set_ids:
        result[sid] = found_ids.get(sid, False)
        
    return jsonify(result)

@api_bp.route('/favorite_all', methods=['POST'])
@login_required
def favorite_all():
    data = request.get_json()
    drug_name = data.get('drug_name')
    batch_id_search = data.get('batch_id_search')
    import_id = data.get('import_id')
    project_id = data.get('project_id')
    new_project_name = data.get('new_project_name')

    target_project = None
    if new_project_name:
        if Project.query.filter_by(owner_id=current_user.id).count() >= 100:
             return jsonify({'error': 'Project limit reached (Max 100). Cannot create new project.'}), 403
        
        target_project = Project(title=new_project_name, owner_id=current_user.id)
        db.session.add(target_project)
        db.session.commit()
    elif project_id:
        target_project = Project.query.get(project_id)
        if not target_project or (target_project.owner_id != current_user.id and current_user not in target_project.members):
            return jsonify({'error': 'Invalid project or unauthorized access.'}), 403
    else:
        target_project = Project.query.filter_by(owner_id=current_user.id, title="Favorite").first()
        if not target_project:
             target_project = Project(title="Favorite", owner_id=current_user.id, display_order=0)
             db.session.add(target_project)
             db.session.commit()

    labels = []
    MAX_FETCH = 100
    
    if batch_id_search:
        ids_list = [sid.strip() for sid in batch_id_search.split(',') if sid.strip()]
        labels, _ = find_labels_by_set_ids(ids_list, skip=0, limit=MAX_FETCH)
    elif import_id:
        import_path = os.path.join(Config.UPLOAD_FOLDER, f"import_{import_id}.json")
        if os.path.exists(import_path):
            with open(import_path, 'r', encoding='utf-8') as f:
                labels = json.load(f)
    elif drug_name:
        labels, _ = find_labels(drug_name, skip=0, limit=MAX_FETCH)
    
    if not labels:
         return jsonify({'success': True, 'added_count': 0, 'project_title': target_project.title})

    added_count = 0
    for label in labels:
        set_id = label['set_id']
        # Check if already exists IN THIS PROJECT (regardless of user)
        existing = Favorite.query.filter_by(set_id=set_id, project_id=target_project.id).first()
        if not existing:
            brand = label.get('brand_name', 'n/a')
            manufacturer = label.get('manufacturer_name', 'n/a')

            # XML Fallback for favorite_all
            if brand.lower() in ['n/a', 'unknown drug'] or manufacturer.lower() in ['n/a', 'unknown manufacturer']:
                from dashboard.services.fda_client import get_label_xml
                from dashboard.services.xml_handler import extract_metadata_from_xml
                xml_raw = get_label_xml(set_id)
                if xml_raw:
                    xml_meta = extract_metadata_from_xml(xml_raw)
                    if xml_meta:
                        if brand.lower() in ['n/a', 'unknown drug'] and xml_meta.get('brand_name') and xml_meta.get('brand_name') != 'Unknown Drug':
                            brand = xml_meta['brand_name']
                        if manufacturer.lower() in ['n/a', 'unknown manufacturer'] and xml_meta.get('manufacturer_name') and xml_meta.get('manufacturer_name') != 'Unknown Manufacturer':
                            manufacturer = xml_meta['manufacturer_name']

            new_fav = Favorite(
                user_id=current_user.id,
                project_id=target_project.id,
                set_id=set_id,
                brand_name=brand,
                generic_name=label.get('generic_name', 'n/a'),
                manufacturer_name=manufacturer,
                market_category=label.get('market_category', 'n/a'),
                application_number=label.get('application_number', 'n/a'),
                ndc=label.get('ndc', 'n/a'),
                effective_time=label.get('effective_time', 'n/a'),
                # New columns
                active_ingredients=label.get('active_ingredients', 'n/a'),
                labeling_type=label.get('labeling_type', 'n/a'),
                dosage_forms=label.get('dosage_forms', 'n/a'),
                routes=label.get('routes', 'n/a'),
                epc=label.get('epc', 'n/a'),
                fdalabel_link=f"https://nctr-crs.fda.gov/fdalabel/ui/search/spl/{set_id}",
                dailymed_spl_link=f"https://dailymed.nlm.nih.gov/dailymed/lookup.cfm?setid={set_id}",
                dailymed_pdf_link=f"https://dailymed.nlm.nih.gov/dailymed/getpdf.cfm?setid={set_id}",
                product_type=label.get('product_type', 'n/a'),
                label_format=label.get('label_format', 'n/a'),
                source=label.get('source', 'n/a')
            )
            db.session.add(new_fav)
            added_count += 1
            
    db.session.commit()
    
    return jsonify({
        'success': True, 
        'added_count': added_count, 
        'project_title': target_project.title
    })

@api_bp.route('/my_labelings', methods=['GET'])
@login_required
def get_my_labelings():
    try:
        # Implement logic to fetch user's projects or labelings
        # For now, return a simple response
        return jsonify({'success': True, 'message': 'My Labelings endpoint'})
    except Exception as e:
        logger.error(f"Error in get_my_labelings: {e}")
        return jsonify({'error': 'Failed to fetch my labelings'}), 500

# --- MedDRA Search ---
@api_bp.route('/meddra/search', methods=['GET'])
def meddra_search():
    query = request.args.get('q', '')
    if len(query) < 2:
        return jsonify([])
    
    try:
        # Search for PTs starting with or containing the query
        results = MeddraPT.query.filter(
            MeddraPT.pt_name.ilike(f'%{query}%')
        ).order_by(MeddraPT.pt_name).limit(10).all()
        
        return jsonify([pt.pt_name for pt in results])
    except Exception as e:
        logger.error(f"Error in meddra_search: {e}")
        return jsonify({'error': 'Search failed'}), 500

# --- AE Profile Reports ---

def extract_ngram(text, target, n_prev=5, n_after=5):
    """Extracts a word-based n-gram around the target term."""
    # Find the target term ignoring case
    match = re.search(re.escape(target), text, re.IGNORECASE)
    if not match:
        return None
    
    start_idx, end_idx = match.span()
    
    # Get text before and after
    pre_text = text[:start_idx]
    post_text = text[end_idx:]
    
    # Split into words (handling multiple spaces/newlines)
    pre_words = pre_text.split()
    post_words = post_text.split()
    
    # Take the last N from pre and first N from post
    context_pre = " ".join(pre_words[-n_prev:]) if pre_words else ""
    context_post = " ".join(post_words[:n_after]) if post_words else ""
    
    # Reconstruct the exact matched term to preserve its casing/punctuation if needed,
    # or just use the target. Let's use the actual text from the document.
    actual_match = text[start_idx:end_idx]
    
    ngram = f"{context_pre} {actual_match} {context_post}".strip()
    return ngram

def run_ae_report_generation(app, report_id, project_id, target_pt):
    with app.app_context():
        report = ProjectAeReport.query.get(report_id)
        if not report:
            return

        try:
            report.status = 'processing'
            db.session.commit()

            # 1. Get all labels in project
            favorites = Favorite.query.filter_by(project_id=project_id).all()
            total_favs = len(favorites)
            report.total_labels = total_favs
            db.session.commit()

            if total_favs == 0:
                report.status = 'completed'
                report.progress = 100
                report.completed_at = datetime.utcnow()
                db.session.commit()
                return

            # Target sections for scan
            target_sections = {
                '34066-1': 'Boxed Warning',
                '34070-3': 'Contraindications',
                '34071-1': 'Warnings and Precautions',
                '43685-7': 'Warnings and Precautions',
                '34084-4': 'Adverse Reactions'
            }

            details_map = [] # List of dicts to store detail data before saving
            unique_drugs = {} # name -> list of indices in details_map

            # --- PHASE 1: Labeling Scan (0-50%) ---
            for i, fav in enumerate(favorites):
                db.session.refresh(report)
                if report.status != 'processing':
                    return

                # A. Text Matching
                xml_content = get_label_xml(fav.set_id)
                found_sections = []
                is_labeled = False
                
                if xml_content:
                    try:
                        ns = {'v3': 'urn:hl7-org:v3'}
                        root = ET.fromstring(xml_content.encode('ascii', 'ignore').decode('ascii'))
                        
                        for section in root.findall(".//v3:section", ns):
                            code_el = section.find("v3:code", ns)
                            if code_el is not None:
                                code_val = code_el.get('code')
                                if code_val in target_sections:
                                    section_name = target_sections[code_val]
                                    text_content = "".join(section.itertext()).strip()
                                    
                                    matches = re.finditer(re.escape(target_pt), text_content, re.IGNORECASE)
                                    for m in matches:
                                        idx = m.start()
                                        start = max(0, idx - 400)
                                        end = min(len(text_content), idx + len(target_pt) + 400)
                                        window_text = text_content[start:end]
                                        ngram = extract_ngram(window_text, target_pt, n_prev=5, n_after=5)
                                        
                                        if ngram:
                                            is_dup = any(s['snippet'] == ngram for s in found_sections)
                                            if not is_dup:
                                                is_labeled = True
                                                found_sections.append({'section': section_name, 'snippet': ngram})
                    except Exception as e:
                        logger.error(f"Error parsing XML for report {report_id}, label {fav.set_id}: {e}")

                drug_name = (fav.generic_name or fav.brand_name or "").split(',')[0].strip()
                if not drug_name or drug_name == 'N/A':
                    drug_name = "Unknown"

                detail_data = {
                    'set_id': fav.set_id,
                    'brand_name': fav.brand_name,
                    'generic_name': fav.generic_name,
                    'is_labeled': is_labeled,
                    'found_sections': json.dumps(found_sections),
                    'drug_name_for_api': drug_name
                }
                details_map.append(detail_data)
                
                if drug_name != "Unknown":
                    if drug_name not in unique_drugs:
                        unique_drugs[drug_name] = []
                    unique_drugs[drug_name].append(len(details_map) - 1)

                # Update progress (0-50%)
                report.progress = int(((i + 1) / total_favs) * 50)
                db.session.commit()

            # --- PHASE 2: openFDA Scan (51-100%) ---
            unique_drug_names = [d for d in unique_drugs.keys() if d != "Unknown"]
            total_unique = len(unique_drug_names)
            
            # Prepare date ranges
            now = datetime.now()
            date_today = now.strftime('%Y%m%d')
            date_1yr_ago = (now - timedelta(days=365)).strftime('%Y%m%d')
            date_5yr_ago = (now - timedelta(days=365*5)).strftime('%Y%m%d')

            for i, drug_name in enumerate(unique_drug_names):
                db.session.refresh(report)
                if report.status != 'processing':
                    return

                counts = {'all': 0, '1yr': 0, '5yr': 0}
                try:
                    base_url = "https://api.fda.gov/drug/event.json"
                    search_base = f'(patient.drug.openfda.brand_name:"{drug_name}" OR patient.drug.openfda.generic_name:"{drug_name}") AND patient.reaction.reactionmeddrapt.exact:"{target_pt}"'
                    
                    # 1. All counts
                    params = {'search': search_base}
                    if Config.OPENFDA_API_KEY: params['api_key'] = Config.OPENFDA_API_KEY
                    resp = requests.get(base_url, params=params)
                    if resp.status_code == 200:
                        counts['all'] = resp.json().get('meta', {}).get('results', {}).get('total', 0)
                    
                    if counts['all'] > 0:
                        # 2. Last 1 year
                        params_1y = {'search': f"{search_base} AND receivedate:[{date_1yr_ago} TO {date_today}]"}
                        if Config.OPENFDA_API_KEY: params_1y['api_key'] = Config.OPENFDA_API_KEY
                        resp_1y = requests.get(base_url, params=params_1y)
                        if resp_1y.status_code == 200:
                            counts['1yr'] = resp_1y.json().get('meta', {}).get('results', {}).get('total', 0)
                            
                        # 3. Last 5 years
                        params_5y = {'search': f"{search_base} AND receivedate:[{date_5yr_ago} TO {date_today}]"}
                        if Config.OPENFDA_API_KEY: params_5y['api_key'] = Config.OPENFDA_API_KEY
                        resp_5y = requests.get(base_url, params=params_5y)
                        if resp_5y.status_code == 200:
                            counts['5yr'] = resp_5y.json().get('meta', {}).get('results', {}).get('total', 0)

                except Exception as e:
                    logger.error(f"FAERS error for report {report_id}, drug {drug_name}: {e}")

                # Map back to all details sharing this drug name
                for idx in unique_drugs[drug_name]:
                    details_map[idx]['faers_count'] = counts['all']
                    details_map[idx]['faers_1yr_count'] = counts['1yr']
                    details_map[idx]['faers_5yr_count'] = counts['5yr']

                # Update progress (51-100%)
                if total_unique > 0:
                    report.progress = 50 + int(((i + 1) / total_unique) * 50)
                else:
                    report.progress = 100
                db.session.commit()

            # Final Save to DB
            for d_data in details_map:
                detail = ProjectAeReportDetail(
                    report_id=report_id,
                    set_id=d_data['set_id'],
                    brand_name=d_data['brand_name'],
                    generic_name=d_data['generic_name'],
                    is_labeled=d_data['is_labeled'],
                    found_sections=d_data['found_sections'],
                    faers_count=d_data.get('faers_count', 0),
                    faers_1yr_count=d_data.get('faers_1yr_count', 0),
                    faers_5yr_count=d_data.get('faers_5yr_count', 0)
                )
                db.session.add(detail)

            report.status = 'completed'
            report.completed_at = datetime.utcnow()
            report.progress = 100
            db.session.commit()

        except Exception as e:
            logger.error(f"AE Report Generation Failed: {e}")
            report.status = 'failed'
            db.session.commit()

        except Exception as e:
            logger.error(f"AE Report Generation Failed: {e}")
            report.status = 'failed'
            db.session.commit()

@api_bp.route('/ae_report/generate', methods=['POST'])
@login_required
def generate_ae_report():
    data = request.get_json()
    project_id = data.get('project_id')
    target_pt = data.get('target_pt')

    if not project_id or not target_pt:
        return jsonify({'error': 'Missing project_id or target_pt'}), 400

    project = Project.query.get(project_id)
    if not project or (project.owner_id != current_user.id and current_user not in project.members):
        return jsonify({'error': 'Unauthorized or project not found'}), 403

    # Create report entry
    new_report = ProjectAeReport(
        project_id=project_id,
        target_pt=target_pt,
        status='pending'
    )
    db.session.add(new_report)
    db.session.commit()

    # Start background task
    from flask import current_app
    app = current_app._get_current_object()
    thread = threading.Thread(target=run_ae_report_generation, args=(app, new_report.id, project_id, target_pt))
    thread.start()

    return jsonify({'success': True, 'report_id': new_report.id})

@api_bp.route('/ae_report/reanalyze/<int:report_id>', methods=['POST'])
@login_required
def reanalyze_ae_report(report_id):
    try:
        report = ProjectAeReport.query.get_or_404(report_id)
        # Check permissions
        if report.project.owner_id != current_user.id and current_user not in report.project.members:
            return jsonify({'error': 'Unauthorized'}), 403

        # Reset report status and progress
        report.status = 'pending'
        report.progress = 0
        report.processed_labels = 0
        report.completed_at = None
        
        # Delete previous details efficiently
        ProjectAeReportDetail.query.filter_by(report_id=report.id).delete(synchronize_session=False)
            
        db.session.commit()
        logger.info(f"Re-analyzing report {report_id} for project {report.project_id}")

        # Restart background task
        from flask import current_app
        app = current_app._get_current_object()
        thread = threading.Thread(target=run_ae_report_generation, args=(app, report.id, report.project_id, report.target_pt))
        thread.start()

        return jsonify({'success': True, 'report_id': report.id})
    except Exception as e:
        logger.error(f"Error in reanalyze_ae_report: {e}")
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/ae_report/status/<int:report_id>')
@login_required
def get_ae_report_status(report_id):
    report = ProjectAeReport.query.get_or_404(report_id)
    # Check permissions (through project)
    if report.project.owner_id != current_user.id and current_user not in report.project.members:
        return jsonify({'error': 'Unauthorized'}), 403

    return jsonify({
        'id': report.id,
        'status': report.status,
        'progress': report.progress,
        'processed': report.processed_labels,
        'total': report.total_labels,
        'target_pt': report.target_pt
    })

@api_bp.route('/ae_report/active_tasks')
@login_required
def get_active_ae_tasks():
    # Get all projects where user is owner or member
    user_projects = Project.query.filter(
        (Project.owner_id == current_user.id) | 
        (Project.members.contains(current_user))
    ).all()
    project_ids = [p.id for p in user_projects]
    
    active_reports = ProjectAeReport.query.filter(
        ProjectAeReport.project_id.in_(project_ids),
        ProjectAeReport.status.in_(['processing', 'pending'])
    ).all()
    
    return jsonify([{
        'id': r.id,
        'project_id': r.project_id,
        'project_title': r.project.title,
        'target_pt': r.target_pt,
        'progress': r.progress,
        'status': r.status
    } for r in active_reports])

@api_bp.route('/ae_report/list/<int:project_id>')
@login_required
def list_ae_reports(project_id):
    project = Project.query.get_or_404(project_id)
    if project.owner_id != current_user.id and current_user not in project.members:
        return jsonify({'error': 'Unauthorized'}), 403

    reports = ProjectAeReport.query.filter_by(project_id=project_id).order_by(ProjectAeReport.created_at.desc()).all()
    return jsonify([{
        'id': r.id,
        'target_pt': r.target_pt,
        'status': r.status,
        'progress': r.progress,
        'created_at': r.created_at.isoformat()
    } for r in reports])

@api_bp.route('/ae_report/delete/<int:report_id>', methods=['DELETE'])
@login_required
def delete_ae_report(report_id):
    logger.info(f"Attempting to delete report {report_id}...")
    try:
        report = ProjectAeReport.query.get(report_id)
        if not report:
            logger.warning(f"Report {report_id} not found for deletion.")
            return jsonify({'error': 'Report not found'}), 404

        # Check permissions (must be project owner or member)
        logger.info(f"Checking permissions for report {report_id}...")
        if report.project.owner_id != current_user.id and current_user not in report.project.members:
            return jsonify({'error': 'Unauthorized'}), 403

        # Mark as deleted first so any background thread stops
        logger.info(f"Setting status to deleted for report {report_id}...")
        report.status = 'deleted'
        db.session.commit()
        
        # Now delete the actual record (cascades to details)
        logger.info(f"Performing actual deletion of report {report_id}...")
        db.session.delete(report)
        db.session.commit()
        
        logger.info(f"Successfully deleted AE report {report_id}")
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error deleting AE report {report_id}: {str(e)}", exc_info=True)
        db.session.rollback()
        return jsonify({'error': f"Server error: {str(e)}"}), 500

@api_bp.route('/ae_report/detail/<int:report_id>')
@login_required
def get_ae_report_detail(report_id):
    report = ProjectAeReport.query.get_or_404(report_id)
    if report.project.owner_id != current_user.id and current_user not in report.project.members:
        return jsonify({'error': 'Unauthorized'}), 403

    details = ProjectAeReportDetail.query.filter_by(report_id=report_id).all()
    
    # Calculate Frequent Contexts Ranking with Similarity Grouping
    from difflib import SequenceMatcher
    import string
    
    def normalize_for_comparison(text):
        # Remove punctuation and normalize spaces for comparison purposes
        text = text.lower()
        text = text.translate(str.maketrans('', '', string.punctuation))
        return " ".join(text.split())

    def get_similarity(a, b):
        return SequenceMatcher(None, normalize_for_comparison(a), normalize_for_comparison(b)).ratio()

    # Clusters: { "representative_ngram": set([set_id1, set_id2...]) }
    clusters = {}
    SIMILARITY_THRESHOLD = 0.80 # Slightly lower threshold for better grouping

    for d in details:
        if d.found_sections:
            secs = json.loads(d.found_sections)
            for s in secs:
                norm = s['snippet'].strip('. ').strip()
                if not norm: continue
                
                # Check if this matches any existing cluster representative
                matched_rep = None
                for rep in clusters:
                    if get_similarity(norm, rep) >= SIMILARITY_THRESHOLD:
                        matched_rep = rep
                        break
                
                if matched_rep:
                    clusters[matched_rep].add(d.set_id)
                else:
                    # New cluster
                    clusters[norm] = {d.set_id}
    
    # Get top 20 contexts from clusters
    frequent_contexts = []
    # Sort the clusters by their unique document counts
    sorted_clusters = sorted(clusters.items(), key=lambda x: len(x[1]), reverse=True)
    
    for snippet, set_ids in sorted_clusters[:20]:
        frequent_contexts.append({
            'snippet': snippet,
            'count': len(set_ids),
            'set_ids': list(set_ids)
        })

    return jsonify({
        'report': {
            'id': report.id,
            'target_pt': report.target_pt,
            'project_title': report.project.title,
            'created_at': report.created_at.isoformat(),
            'status': report.status
        },
        'frequent_contexts': frequent_contexts,
        'results': [{
            'set_id': d.set_id,
            'brand_name': d.brand_name,
            'generic_name': d.generic_name,
            'is_labeled': d.is_labeled,
            'found_sections': json.loads(d.found_sections) if d.found_sections else [],
            'faers_count': d.faers_count or 0,
            'faers_1yr_count': d.faers_1yr_count or 0,
            'faers_5yr_count': d.faers_5yr_count or 0
        } for d in details]
    })

@api_bp.route('/ae_report/export_json/<int:report_id>')
@login_required
def export_ae_report_json(report_id):
    report = ProjectAeReport.query.get_or_404(report_id)
    if report.project.owner_id != current_user.id and current_user not in report.project.members:
        return jsonify({'error': 'Unauthorized'}), 403

    details = ProjectAeReportDetail.query.filter_by(report_id=report_id).all()
    
    # Aggregate contexts for the export with similarity clustering
    from difflib import SequenceMatcher
    import string
    
    def normalize_for_comparison(text):
        # Remove punctuation and normalize spaces for comparison purposes
        text = text.lower()
        text = text.translate(str.maketrans('', '', string.punctuation))
        return " ".join(text.split())

    def get_similarity(a, b):
        return SequenceMatcher(None, normalize_for_comparison(a), normalize_for_comparison(b)).ratio()
        
    clusters = {} # snippet -> set of set_ids
    label_data = []
    SIMILARITY_THRESHOLD = 0.80
    
    for d in details:
        found_secs = json.loads(d.found_sections) if d.found_sections else []
        for s in found_secs:
            norm = s['snippet'].strip('. ').strip()
            if not norm: continue
            
            # Grouping logic
            matched_rep = None
            for rep in clusters:
                if get_similarity(norm, rep) >= SIMILARITY_THRESHOLD:
                    matched_rep = rep
                    break
            
            if matched_rep:
                clusters[matched_rep].add(d.set_id)
            else:
                clusters[norm] = {d.set_id}
                
        label_data.append({
            'brand_name': d.brand_name,
            'generic_name': d.generic_name,
            'set_id': d.set_id,
            'is_labeled': d.is_labeled,
            'mentions': found_secs,
            'faers_metrics': {
                'total_count': d.faers_count,
                'last_1yr_count': d.faers_1yr_count,
                'last_5yr_count': d.faers_5yr_count
            }
        })

    # Prepare AI prompt
    ai_prompt = (
        f"The following data represents a safety analysis for the MedDRA Preferred Term: '{report.target_pt}'. "
        f"It was generated from project '{report.project.title}'. "
        "The 'frequent_contexts' section ranks the most common phrasing found in FDA labeling for this term, "
        "using a similarity clustering algorithm (threshold 0.80) to group near-identical phrasing together, "
        "ignoring punctuation and whitespace during comparison. "
        "The 'label_details' section provides document-level findings, including specific snippets and FAERS (FDA Adverse Event Reporting System) counts. "
        "Analyze this data to identify patterns in how this adverse event is described across different drugs, "
        "correlate labeling mentions with FAERS counts, and summarize the typical clinical context (e.g., severity, population, or co-medications) "
        "associated with this event."
    )

    # Sort clusters by document frequency
    sorted_contexts = sorted(clusters.items(), key=lambda x: len(x[1]), reverse=True)

    # Calculate Unique Drug Stats for Summary
    unique_drugs = {}
    for d in details:
        name = (d.generic_name or d.brand_name or "Unknown").split(',')[0].strip()
        if name not in unique_drugs:
            unique_drugs[name] = {
                'all': d.faers_count or 0,
                '1yr': d.faers_1yr_count or 0,
                '5yr': d.faers_5yr_count or 0
            }

    export_obj = {
        'metadata': {
            'report_id': report.id,
            'target_pt': report.target_pt,
            'project_title': report.project.title,
            'generated_at': datetime.utcnow().isoformat(),
            'total_labels_analyzed': len(details)
        },
        'ai_instructions': ai_prompt,
        'summary_statistics': {
            'labeled_presence_count': len([d for d in details if d.is_labeled]),
            'total_faers_reports': sum(v['all'] for v in unique_drugs.values()),
            'last_1yr_total': sum(v['1yr'] for v in unique_drugs.values()),
            'last_5yr_total': sum(v['5yr'] for v in unique_drugs.values())
        },
        'frequent_contexts': [
            {'snippet': s, 'document_frequency': len(set_ids)} 
            for s, set_ids in sorted_contexts[:50]
        ],
        'label_details': label_data
    }

    return jsonify(export_obj)

@api_bp.route('/ae_report/export/<int:report_id>')
@login_required
def export_ae_report(report_id):
    report = ProjectAeReport.query.get_or_404(report_id)
    if report.project.owner_id != current_user.id and current_user not in report.project.members:
        return jsonify({'error': 'Unauthorized'}), 403

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from flask import send_file

    details = ProjectAeReportDetail.query.filter_by(report_id=report_id).all()
    
    wb = Workbook()
    
    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    
    ws1['A1'] = "AE Profile Report Summary"
    ws1['A1'].font = Font(size=14, bold=True)
    
    # Calculate Unique Drug Stats for Summary
    unique_drugs = {}
    for d in details:
        name = (d.generic_name or d.brand_name or "Unknown").split(',')[0].strip()
        if name not in unique_drugs:
            unique_drugs[name] = {
                'all': d.faers_count or 0,
                '1yr': d.faers_1yr_count or 0,
                '5yr': d.faers_5yr_count or 0
            }

    summary_data = [
        ["Target PT", report.target_pt],
        ["Project", report.project.title],
        ["Report Created", report.created_at.strftime('%Y-%m-%d %H:%M')],
        ["Total Labels", len(details)],
        ["Labeled Labels", len([d for d in details if d.is_labeled])],
        ["Total FAERS Reports (All)", sum(v['all'] for v in unique_drugs.values())],
        ["FAERS Reports (Last 5 Years)", sum(v['5yr'] for v in unique_drugs.values())],
        ["FAERS Reports (Last 1 Year)", sum(v['1yr'] for v in unique_drugs.values())]
    ]
    
    for r_idx, row in enumerate(summary_data, 3):
        ws1.cell(row=r_idx, column=1, value=row[0]).font = Font(bold=True)
        ws1.cell(row=r_idx, column=2, value=row[1])

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 35

    # Sheet 2: Data
    ws2 = wb.create_sheet("Labeling Analysis")
    headers = ["Drug Name", "Generic Name", "Labeled?", "Sections Mentioned", "All Counts", "Last 5y", "Last 1y", "Set ID"]
    ws2.append(headers)
    
    # Style header
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for d in details:
        found_secs = json.loads(d.found_sections) if d.found_sections else []
        sec_str = ", ".join([s['section'] for s in found_secs])
        ws2.append([
            d.brand_name,
            d.generic_name,
            "Yes" if d.is_labeled else "No",
            sec_str,
            d.faers_count,
            d.faers_5yr_count,
            d.faers_1yr_count,
            d.set_id
        ])

    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws2.column_dimensions[column].width = min(40, max_length + 2)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"AE_Report_{report.target_pt.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

# --- FAERS & Assessment ---
@api_bp.route('/faers/<path:drug_name>')
def api_faers_data(drug_name):
    limit = request.args.get('limit', default=20, type=int)
    data = get_faers_data(drug_name, limit=limit)
    if not data:
        return jsonify({'error': 'Could not fetch data'}), 500
    
    # Enrich with MedDRA Data
    if 'reactions' in data:
        data['reactions'] = enrich_faers_with_meddra(data['reactions'])
    if 'reactions_serious' in data:
        data['reactions_serious'] = enrich_faers_with_meddra(data['reactions_serious'])
    if 'reactions_non_serious' in data:
        data['reactions_non_serious'] = enrich_faers_with_meddra(data['reactions_non_serious'])

    return jsonify(data)

@api_bp.route('/faers/trends', methods=['POST'])
def api_faers_trends():
    data = request.get_json()
    drug_name = data.get('drug_name')
    terms = data.get('terms', [])

    if not drug_name or not terms:
        return jsonify({'error': 'Missing drug_name or terms'}), 400

    base_url = "https://api.fda.gov/drug/event.json"
    search_term = f'(patient.drug.openfda.brand_name:"{drug_name}" OR patient.drug.openfda.generic_name:"{drug_name}")'
    
    current_year = datetime.now().year
    start_date = f"{current_year - 5}0101"
    end_date = f"{current_year}1231"
    
    trends = {}

    for term in terms:
        try:
            term_search = f'{search_term} AND patient.reaction.reactionmeddrapt.exact:"{term}" AND receiptdate:[{start_date} TO {end_date}]'
            params = {
                'search': term_search,
                'count': 'receiptdate'
            }
            if Config.OPENFDA_API_KEY:
                params['api_key'] = Config.OPENFDA_API_KEY
            
            resp = requests.get(base_url, params=params)
            if resp.status_code == 200:
                raw_dates = resp.json().get('results', [])
                raw_dates.sort(key=lambda x: x.get('time', ''))
                trends[term] = raw_dates
            else:
                trends[term] = []
        except Exception as e:
            logger.error(f"Error fetching trend for {term}: {e}")
            trends[term] = []

    return jsonify({'trends': trends})

@api_bp.route('/faers/emerging', methods=['POST'])
def api_faers_emerging():
    """
    Identifies 'emerging' AEs: terms present in the last 5 years but absent 6-10 years ago.
    Also scans the current label (set_id) for these terms.
    """
    data = request.get_json()
    drug_name = data.get('drug_name')
    set_id = data.get('set_id')
    if not drug_name:
        return jsonify({'error': 'Missing drug_name'}), 400

    clean_name = re.split(r'[,;]', drug_name)[0].strip()
    base_url = "https://api.fda.gov/drug/event.json"
    
    # We use openfda.brand_name OR openfda.generic_name for precision
    search_term = f'(patient.drug.openfda.brand_name:"{clean_name}" OR patient.drug.openfda.generic_name:"{clean_name}")'

    # Calculate date ranges: Recent (5y) and Previous (6-10y ago)
    now = datetime.now()
    p1_end = now.strftime('%Y%m%d')
    p1_start = (now - timedelta(days=5*365)).strftime('%Y%m%d')
    p2_end = (now - timedelta(days=5*365 + 1)).strftime('%Y%m%d')
    p2_start = (now - timedelta(days=10*365)).strftime('%Y%m%d')

    try:
        def fetch_counts(start, end):
            query = f'{search_term} AND receivedate:[{start} TO {end}]'
            params = {
                'search': query,
                'count': 'patient.reaction.reactionmeddrapt.exact',
                'limit': 1000
            }
            if Config.OPENFDA_API_KEY:
                params['api_key'] = Config.OPENFDA_API_KEY
            
            resp = requests.get(base_url, params=params, timeout=15)
            if resp.status_code == 200:
                return {r['term']: r['count'] for r in resp.json().get('results', [])}
            elif resp.status_code == 404:
                return {}
            else:
                logger.warning(f"openFDA error {resp.status_code} for period {start}-{end}")
                return {}

        counts_recent = fetch_counts(p1_start, p1_end)
        counts_prev = fetch_counts(p2_start, p2_end)

        emerging = []
        for term, count in counts_recent.items():
            if term not in counts_prev:
                emerging.append({'term': term, 'count': count, 'prev_count': 0, 'label_matches': []})

        # Sort by count desc
        emerging.sort(key=lambda x: x['count'], reverse=True)

        # Enrich with MedDRA (SOC, HLT info)
        if emerging:
            emerging = enrich_faers_with_meddra(emerging)
            
        # LABELING SCAN for Emerging AEs
        if set_id and emerging:
            xml_content = get_label_xml(set_id)
            if xml_content:
                try:
                    ns = {'v3': 'urn:hl7-org:v3'}
                    root = ET.fromstring(xml_content.encode('ascii', 'ignore').decode('ascii'))
                    target_sections = {
                        '34066-1': 'Boxed Warning',
                        '34070-3': 'Contraindications',
                        '34071-1': 'Warnings and Precautions',
                        '43685-7': 'Warnings and Precautions',
                        '34084-4': 'Adverse Reactions'
                    }
                    
                    # Extract text from relevant sections once
                    sections_data = []
                    for section in root.findall(".//v3:section", ns):
                        code_el = section.find("v3:code", ns)
                        if code_el is not None and code_el.get('code') in target_sections:
                            title_el = section.find("v3:title", ns)
                            title = "".join(title_el.itertext()).strip() if title_el is not None else target_sections[code_el.get('code')]
                            text = "".join(section.itertext()).strip()
                            sections_data.append({'title': title, 'text': text})
                    
                    # Match each emerging AE
                    for ae in emerging:
                        term = ae['term']
                        for sec in sections_data:
                            match = re.search(re.escape(term), sec['text'], re.IGNORECASE)
                            if match:
                                start_idx = match.start()
                                end_idx = match.end()
                                context_start = max(0, start_idx - 50)
                                context_end = min(len(sec['text']), end_idx + 100)
                                snippet = sec['text'][context_start:context_end]
                                # Add simple bolding
                                snippet = snippet.replace(sec['text'][start_idx:end_idx], f"**{sec['text'][start_idx:end_idx]}**")
                                
                                ae['label_matches'].append({
                                    'section': sec['title'],
                                    'snippet': snippet
                                })
                except Exception as xml_err:
                    logger.error(f"Error scanning XML for emerging AEs: {xml_err}")

        return jsonify({
            'emerging': emerging,
            'metadata': {
                'drug': clean_name,
                'recent_period': [p1_start, p1_end],
                'previous_period': [p2_start, p2_end]
            }
        })

    except Exception as e:
        logger.error(f"Error in api_faers_emerging: {e}")
        return jsonify({'error': str(e)}), 500

@api_bp.route('/faers/ai_results', methods=['GET'])
def api_faers_ai_results():
    set_id = request.args.get('set_id')
    drug_name = request.args.get('drug_name')
    if not set_id or not drug_name:
        return jsonify({'error': 'Missing parameters'}), 400
    
    assessment = AeAiAssessment.query.filter_by(set_id=set_id, drug_name=drug_name).first()
    if assessment:
        return jsonify({
            'results': json.loads(assessment.result_json),
            'timestamp': assessment.timestamp.isoformat(),
            'min_count': assessment.min_count
        })
    return jsonify({'results': None})

@api_bp.route('/faers/ai_rematch', methods=['POST'])
def api_faers_ai_rematch():
    data = request.get_json()
    set_id = data.get('set_id')
    drug_name = data.get('drug_name')
    terms = data.get('terms', []) # List of {term, count}
    min_count = data.get('min_count', 10)

    if not set_id or not drug_name or not terms:
        return jsonify({'error': 'Missing parameters'}), 400

    # Filter terms by min_count
    filtered_terms = [t for t in terms if t['count'] >= min_count]
    if not filtered_terms:
        return jsonify({'error': f'No terms found with count >= {min_count}'}), 400

    xml_content = get_label_xml(set_id)
    if not xml_content:
        return jsonify({'error': 'Label not found'}), 404

    try:
        # Extract text for AI
        root = ET.fromstring(xml_content.encode('ascii', 'ignore').decode('ascii'))
        text_content = " ".join(root.itertext())
        
        # Limit text size if necessary, but label is usually okay
        # For efficiency, we only send terms AI hasn't seen or that were undocumented
        term_list_str = ", ".join([t['term'] for t in filtered_terms])

        prompt = f"""
        Analyze the provided drug labeling text for the drug "{drug_name}".
        We have a list of adverse event (AE) terms (MedDRA Preferred Terms) reported in FAERS that were NOT found via direct string matching in the label.
        
        TASK: For each term, determine if it is SEMANTICALLY mentioned or related in the labeling (e.g., as a different synonym, a broader category, or mentioned in a specific clinical context).
        
        AE TERMS TO CHECK: {term_list_str}
        
        OUTPUT FORMAT: You MUST return a strict JSON array of objects. No other text.
        Example format:
        [
          {{
            "term": "Term Name",
            "found": true/false,
            "section": "Section Title if found",
            "explanation": "Brief semantic explanation"
          }}
        ]
        """

        user_obj = current_user._get_current_object() if current_user.is_authenticated else None
        # We use ai_chat logic but with a specialized prompt
        ai_response = chat_with_document(user_obj, prompt, [], xml_content, chat_type='general')
        
        # Parse JSON from AI response
        try:
            # Clean possible markdown
            cleaned_json = ai_response.replace('```json', '').replace('```', '').strip()
            result_list = json.loads(cleaned_json)
        except Exception as parse_err:
            logger.error(f"AI JSON Parse Error: {parse_err}. Response: {ai_response}")
            # Try to find JSON block with regex
            match = re.search(r'\[\s*\{.*\}\s*\]', ai_response, re.DOTALL)
            if match:
                result_list = json.loads(match.group(0))
            else:
                return jsonify({'error': 'AI failed to return valid JSON format', 'raw': ai_response}), 500

        # Save to DB (Upsert)
        existing = AeAiAssessment.query.filter_by(set_id=set_id, drug_name=drug_name).first()
        if existing:
            existing.result_json = json.dumps(result_list)
            existing.min_count = min_count
            existing.timestamp = datetime.utcnow()
        else:
            new_assessment = AeAiAssessment(
                set_id=set_id,
                drug_name=drug_name,
                result_json=json.dumps(result_list),
                min_count=min_count
            )
            db.session.add(new_assessment)
        db.session.commit()

        return jsonify({
            'results': result_list,
            'timestamp': datetime.utcnow().isoformat(),
            'min_count': min_count
        })

    except Exception as e:
        logger.error(f"Error in api_faers_ai_rematch: {e}")
        return jsonify({'error': str(e)}), 500

def generic_assessment_route(set_id, assessment_model, pt_terms, prompt, keyword_check_fn):
    # Check existing
    try:
        # Check ToxAgent FIRST (Consolidated Table)
        tox_agent = ToxAgent.query.filter_by(set_id=set_id, current='Yes').first()
        
        report_field_map = {
            DiliAssessment: 'dili_report',
            DictAssessment: 'dict_report',
            DiriAssessment: 'diri_report'
        }
        
        field_name = report_field_map.get(assessment_model)
        existing_report = None
        timestamp = None
        
        def is_complete(report):
            if not report: return False
            if '<!-- AI response did not contain' in report: return False
            if '<div' in report or ('<!--' in report and 'evidence found' in report):
                return True
            return False

        if tox_agent and field_name:
            report_val = getattr(tox_agent, field_name)
            if is_complete(report_val):
                existing_report = report_val
                timestamp = tox_agent.last_updated.isoformat()
        
        # Fallback to individual tables if not found in ToxAgent
        if not existing_report:
            assessment = assessment_model.query.filter_by(set_id=set_id).first()
            if assessment and is_complete(assessment.report_content):
                existing_report = assessment.report_content
                timestamp = assessment.timestamp.isoformat()
            
    except Exception as e:
        logger.error(f"Database error in generic_assessment_route: {e}")
        existing_report = None
        timestamp = None
    
    # FAERS data check
    faers_data = []
    faers_error = None
    meta = get_label_metadata(set_id)
    if meta:
        brand_name = meta.get('brand_name', '').split(',')[0].strip()
        generic_name = meta.get('generic_name', '').split(',')[0].strip()

        # Use first non-N/A name
        drug_name = brand_name if (brand_name and brand_name != 'N/A') else generic_name

        if drug_name and drug_name != 'N/A':
            base_url = "https://api.fda.gov/drug/event.json"

            # Improved search: search both brand and generic in medicinalproduct field
            # and generic_name field if possible.
            # Simplified: use the provided keyword_check_fn but wrap the query logic
            def enhanced_check_fn(b_name, g_name):
                # Build an OR clause for brand and generic
                name_clause = f'(patient.drug.medicinalproduct:"{b_name}"'
                if g_name and g_name != b_name:
                    name_clause += f' OR patient.drug.medicinalproduct:"{g_name}"'
                name_clause += ')'

                # Extract the reaction part from the original fn (assuming it's after the first AND)
                original_q = keyword_check_fn(b_name)
                reaction_part = original_q.split(' AND ', 1)[1] if ' AND ' in original_q else ""

                return f'{name_clause} AND {reaction_part}'

            search_query = enhanced_check_fn(brand_name, generic_name)
            params = {
                'search': search_query,
                'count': 'patient.reaction.reactionmeddrapt.exact',
                'limit': 1000
            }
            if Config.OPENFDA_API_KEY:
                params['api_key'] = Config.OPENFDA_API_KEY
                
            try:
                resp = requests.get(base_url, params=params, timeout=10)
                if resp.status_code == 200:
                    raw_results = resp.json().get('results', [])
                    faers_data = [item for item in raw_results if item['term'].upper() in pt_terms]
                elif resp.status_code == 404:
                    # openFDA returns 404 for "No results found"
                    faers_data = []
                else:
                    logger.warning(f"FAERS query failed: {resp.status_code}")
                    faers_error = f"API returned status {resp.status_code}"
            except Exception as e:
                logger.error(f"FAERS error: {e}")
                from dashboard.services.fda_client import handle_openfda_error
                faers_error = handle_openfda_error(e)

    return jsonify({
        'faers_data': faers_data, 
        'faers_error': faers_error,
        'existing_assessment': existing_report,
        'assessment_timestamp': timestamp
    })

def run_assessment_logic(set_id, assessment_model, prompt):
    xml_content = get_label_xml(set_id)
    if not xml_content:
        return jsonify({'error': "Could not retrieve label data"}), 404

    try:
        ns = {'v3': 'urn:hl7-org:v3'}
        xml_string_cleaned = xml_content.encode('ascii', 'ignore').decode('ascii')
        root = ET.fromstring(xml_string_cleaned)

        target_code_map = {
            '34066-1': 'Boxed Warning',
            '34070-3': 'Contraindications',
            '34071-1': 'Warnings and Precautions',
            '43685-7': 'Warnings and Precautions',
            '34084-4': 'Adverse Reactions',
            '34073-7': 'Drug Interactions',
            '43684-0': 'Use in Specific Populations'
        }

        aggregated_parts = []
        processed_ids = set()

        for section in root.findall(".//v3:section", ns):
            code_el = section.find("v3:code", ns)
            if code_el is None:
                continue

            code_val = code_el.get('code')
            if code_val not in target_code_map:
                continue

            sec_id = section.get('ID', str(uuid.uuid4()))
            if sec_id in processed_ids:
                continue
            processed_ids.add(sec_id)

            section_name = target_code_map[code_val]
            text_content = " ".join("".join(section.itertext()).split()).strip()

            if len(text_content) > 10:
                aggregated_parts.append(f"### {section_name}\n{text_content}")

        aggregated_text = "\n\n".join(aggregated_parts)

        if not aggregated_text:
            return jsonify({'assessment_report': "No relevant sections found."})

        try:
            user_obj = current_user._get_current_object() if current_user.is_authenticated else None
            response_text = generate_assessment(user_obj, prompt, aggregated_text)

            html_matches = re.findall(
                r'(<div class="label-section">[\s\S]*?</div>)',
                response_text,
                re.DOTALL
            )

            if html_matches:
                clean_report = "\n".join(html_matches)
            else:
                if '<!-- No DILI evidence found' in response_text or 'No DILI Concern' in response_text:
                    clean_report = '<div class="label-section"><!-- No DILI evidence found in label --><p><strong>Conclusion:</strong> No DILI Concern</p></div>'
                elif '<!-- No DICT evidence found' in response_text or 'No DICT Concern' in response_text:
                    clean_report = '<div class="label-section"><!-- No DICT evidence found in label --><p><strong>Conclusion:</strong> No DICT Concern</p></div>'
                elif '<!-- No DIRI evidence found' in response_text or 'No DIRI Concern' in response_text:
                    clean_report = '<div class="label-section"><!-- No DIRI evidence found in label --><p><strong>Conclusion:</strong> No DIRI Concern</p></div>'
                else:
                    clean_report = '<div class="label-section"><p class="error">AI response did not contain valid HTML report.</p></div>'

            assessment = assessment_model.query.filter_by(set_id=set_id).first()
            if assessment:
                assessment.report_content = clean_report
                assessment.timestamp = datetime.utcnow()
            else:
                assessment = assessment_model(set_id=set_id, report_content=clean_report)
                db.session.add(assessment)

            try:
                meta = extract_metadata_from_xml(xml_content) or {}
                new_eff_time = meta.get('effective_time')

                tox_agent = ToxAgent.query.filter_by(set_id=set_id, current='Yes').first()

                if tox_agent and tox_agent.spl_effective_time != new_eff_time:
                    ToxAgent.query.filter_by(set_id=set_id).update({"current": "No"})
                    tox_agent = None

                if not tox_agent:
                    tox_agent = ToxAgent(
                        set_id=set_id,
                        brand_name=meta.get('brand_name'),
                        generic_name=meta.get('generic_name'),
                        manufacturer=meta.get('manufacturer_name'),
                        spl_effective_time=new_eff_time,
                        is_plr=1 if meta.get('label_format') == 'PLR' else 0,
                        current='Yes',
                        status='pending'
                    )
                    db.session.add(tox_agent)

                report_field_map = {
                    DiliAssessment: 'dili_report',
                    DictAssessment: 'dict_report',
                    DiriAssessment: 'diri_report'
                }
                field_name = report_field_map.get(assessment_model)
                if field_name:
                    setattr(tox_agent, field_name, clean_report)

                if not tox_agent.brand_name:
                    tox_agent.brand_name = meta.get('brand_name')
                if not tox_agent.generic_name:
                    tox_agent.generic_name = meta.get('generic_name')

                tox_agent.last_updated = datetime.utcnow()
                tox_agent.status = 'completed'

            except Exception as tox_err:
                logger.error(f"Failed to update ToxAgent during manual assessment: {tox_err}")

            db.session.commit()
            return jsonify({'assessment_report': clean_report})

        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f"AI Analysis Failed: {str(e)}"}), 500

    except Exception as e:
        logger.error(f"Parsing error: {e}")
        return jsonify({'error': "Error processing request"}), 500


@api_bp.route('/dili/faers/<set_id>')
def api_dili_faers(set_id):
    def check_fn(brand):
        return f'patient.drug.medicinalproduct:"{brand}" AND (patient.reaction.reactionmeddrapt:liver OR patient.reaction.reactionmeddrapt:hepatic)'
    return generic_assessment_route(set_id, DiliAssessment, DILI_PT_TERMS, DILI_prompt, check_fn)

@api_bp.route('/dili/assess/<set_id>')
def api_dili_assess(set_id):
    return run_assessment_logic(set_id, DiliAssessment, DILI_prompt)

@api_bp.route('/dict/faers/<set_id>')
def api_dict_faers(set_id):
    def check_fn(brand):
        return f'patient.drug.medicinalproduct:"{brand}" AND (patient.reaction.reactionmeddrapt:cardiac OR patient.reaction.reactionmeddrapt:heart OR patient.reaction.reactionmeddrapt:myocardial)'
    return generic_assessment_route(set_id, DictAssessment, DICT_PT_TERMS, DICT_prompt, check_fn)

@api_bp.route('/dict/assess/<set_id>')
def api_dict_assess(set_id):
    return run_assessment_logic(set_id, DictAssessment, DICT_prompt)

@api_bp.route('/diri/faers/<set_id>')
def api_diri_faers(set_id):
    def check_fn(brand):
        return f'patient.drug.medicinalproduct:"{brand}" AND (patient.reaction.reactionmeddrapt:renal OR patient.reaction.reactionmeddrapt:kidney)'
    return generic_assessment_route(set_id, DiriAssessment, DIRI_PT_TERMS, DIRI_prompt, check_fn)

@api_bp.route('/diri/assess/<set_id>')
def api_diri_assess(set_id):
    return run_assessment_logic(set_id, DiriAssessment, DIRI_prompt)

@api_bp.route('/pgx/assess/<set_id>')
def api_pgx_assess(set_id):
    try:
        force_refresh = request.args.get('refresh') == 'true'
        user_obj = current_user._get_current_object() if current_user.is_authenticated else None
        result = run_pgx_assessment(set_id, user=user_obj, force_refresh=force_refresh)
        if 'error' in result:
            return jsonify({'error': result['error']}), 500
        return jsonify(result)
    except Exception as e:
        logger.error(f"Critical error in api_pgx_assess: {e}")
        return jsonify({'error': 'Internal Server Error during PGx assessment'}), 500

