from flask import Blueprint, request, redirect, url_for, jsonify, send_from_directory, current_app, send_file
from flask_login import login_required, current_user
import os
import zipfile
import re
import json
from io import BytesIO
from openpyxl import Workbook
from datetime import datetime
from collections import Counter

from database import db, Favorite, Annotation, ToxAgent
from dashboard.services.fdalabel_db import FDALabelDBService
from dashboard.services.xml_handler import (
    parse_spl_xml,
    extract_metadata_from_xml,
)
from dashboard.services.fda_client import (
    get_label_metadata,
    get_label_xml,
    find_labels,
    find_labels_by_set_ids
)

from dashboard.config import Config
import uuid
from openpyxl import load_workbook

main_bp = Blueprint('main', __name__)

@main_bp.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(current_app.root_path, 'static'),
                               'favicon.svg', mimetype='image/svg+xml')

@main_bp.route('/')
def index():
    """ Redirects to the Next.js dashboard. """
    return redirect('/dashboard')

@main_bp.route('/upload_label', methods=['POST'])
def upload_label():
    """
    Handles uploading of an XML or ZIP file containing an SPL label.
    Extracts, validates (PLR/Non-PLR), and stores the label for comparison.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    filename = file.filename.lower()
    if not (filename.endswith('.xml') or filename.endswith('.zip')):
        return jsonify({'error': 'Invalid file format. Please upload .xml or .zip.'}), 400

    # Read current set_ids to validate against
    current_set_ids = request.form.getlist('current_set_ids[]') # AJAX sends array as []
    
    xml_content = None
    
    try:
        if filename.endswith('.zip'):
            with zipfile.ZipFile(file) as z:
                # Find first XML file
                xml_files = [f for f in z.namelist() if f.lower().endswith('.xml')]
                if not xml_files:
                    return jsonify({'error': 'No XML file found in the ZIP archive.'}), 400
                
                with z.open(xml_files[0]) as f:
                    xml_content = f.read().decode('utf-8')
        else:
            xml_content = file.read().decode('utf-8')
            
        if not xml_content:
             return jsonify({'error': 'Empty file content.'}), 400

        # Extract metadata
        meta = extract_metadata_from_xml(xml_content)
        if not meta or not meta.get('set_id'):
            return jsonify({'error': 'Could not parse SPL metadata or Set ID from the XML.'}), 400

        new_label_format = meta.get('label_format')
        new_set_id = meta.get('set_id')

        # VALIDATION: Check against existing labels
        if current_set_ids:
            # We need to check the format of at least one existing label
            existing_meta = get_label_metadata(current_set_ids[0])
            if existing_meta:
                existing_format = existing_meta.get('label_format')
                if existing_format and new_label_format and existing_format != new_label_format:
                     return jsonify({
                         'error': f"Format mismatch: The uploaded label is '{new_label_format}', but you are comparing '{existing_format}' labels. Please upload a compatible label."
                     }), 400

        # Save to local storage
        save_path = os.path.join(Config.UPLOAD_FOLDER, f"{new_set_id}.xml")
        with open(save_path, 'w', encoding='utf-8') as f:
            f.write(xml_content)
            
        return jsonify({'success': True, 'set_id': new_set_id})

    except Exception as e:
        current_app.logger.error(f"Error processing upload: {e}")
        return jsonify({'error': f"Error processing file: {str(e)}"}), 500

@main_bp.route('/import_fdalabel', methods=['POST'])
def import_fdalabel():
    """
    Handles importing drug labels from an FDALabel Excel file.
    Parses the Excel and returns the import ID for the selection page.
    """
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No selected file'}), 400

    if not file.filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'error': 'Invalid file format. Please upload .xlsx'}), 400

    try:
        wb = load_workbook(file, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            return jsonify({'success': False, 'error': 'The Excel file is empty.'}), 400
            
        columns = rows[0]
        # Create a mapping of column name to index, normalizing whitespace
        col_map = {str(name).strip(): i for i, name in enumerate(columns) if name}
        
        # Helper to get value from multiple possible column names
        def get_val_flex(row, candidates):
            for c in candidates:
                if c in col_map:
                    idx = col_map[c]
                    if idx < len(row) and row[idx] is not None:
                        return str(row[idx]).strip()
            return 'n/a'

        # Check for SET ID (essential)
        set_id_candidates = ['SET ID', 'SET_ID', 'Set ID']
        has_set_id = any(c in col_map for c in set_id_candidates)
        
        if not has_set_id:
            return jsonify({'success': False, 'error': 'Missing required column: SET ID'}), 400

        labels = []
        for row in rows[1:]:
            set_id = get_val_flex(row, set_id_candidates)
            if not set_id or set_id.lower() == 'n/a':
                continue
            
            # Map Excel columns to our label structure
            brand_name = get_val_flex(row, ['Trade Name', 'PRODUCT_NAMES', 'trade name', 'Product Names'])
            generic_name = get_val_flex(row, ['Generic/Proper Name(s)', 'PRODUCT_NORMD_GENERIC_NAMES', 'generic name', 'Generic Name'])
            company = get_val_flex(row, ['Company', 'AUTHOR_ORG_NORMD_NAME', 'manufacturer', 'Manufacturer', 'AUTHOR_ORG_NAME'])
            effective_date = get_val_flex(row, ['SPL Effective Date (YYYY/MM/DD)', 'EFF_TIME', 'effective date', 'Effective Date'])
            app_num = get_val_flex(row, ['Application Number(s)', 'APPR_NUM', 'application number', 'Approval Number'])
            labeling_type = get_val_flex(row, ['Labeling Type', 'LABELING_TYPE', 'labeling type', 'Doc Type'])
            ndcs = get_val_flex(row, ['NDC(s)', 'NDC_CODES', 'ndc', 'NDC'])
            marketing_category = get_val_flex(row, ['Marketing Category', 'MARKET_CATEGORIES', 'marketing category'])
            
            # Additional FDALabel specific columns
            dosage_forms = get_val_flex(row, ['Dosage Form(s)', 'DOSAGE_FORMS'])
            routes = get_val_flex(row, ['Route(s) of Administration', 'ROUTES'])
            epc = get_val_flex(row, ['Established Pharmacologic Class(es)', 'EPC'])
            active_ingredients = get_val_flex(row, ['Active Ingredient(s)', 'ACTIVE_INGREDIENTS', 'ACT_INGR_NAMES'])
            
            # Links
            fdalabel_link = get_val_flex(row, ['FDALabel Link'])
            dailymed_spl_link = get_val_flex(row, ['DailyMed SPL Link'])
            dailymed_pdf_link = get_val_flex(row, ['DailyMed PDF Link'])

            # Simplified product type for UI badges
            labeling_type_upper = (labeling_type or "").upper()
            marketing_category_upper = (marketing_category or "").upper()
            prod_type = 'Rx'
            if 'OTC' in labeling_type_upper or 'OTC' in marketing_category_upper:
                prod_type = 'OTC'

            labels.append({
                'set_id': set_id,
                'brand_name': brand_name,
                'generic_name': generic_name,
                'manufacturer_name': company,
                'effective_time': effective_date,
                'label_format': None,
                'labeling_type': labeling_type,
                'application_number': app_num,
                'product_type': prod_type,
                'ndc': ndcs,
                'marketing_category': marketing_category,
                'dosage_forms': dosage_forms,
                'routes': routes,
                'epc': epc,
                'active_ingredients': active_ingredients,
                'fdalabel_link': fdalabel_link,
                'dailymed_spl_link': dailymed_spl_link,
                'dailymed_pdf_link': dailymed_pdf_link,
                'source': 'excel'
            })

        if not labels:
            return jsonify({'success': False, 'error': 'No valid labels found in the Excel file.'}), 400

        # Store in temporary file
        import_id = str(uuid.uuid4())
        import_filename = f"import_{import_id}.json"
        import_path = os.path.join(Config.UPLOAD_FOLDER, import_filename)
        
        with open(import_path, 'w', encoding='utf-8') as f:
            json.dump(labels, f)

        # In a fully Next.js setup, we return the import ID so the frontend can redirect
        return jsonify({
            'success': True, 
            'import_id': import_id,
            'redirect_url': f'/search?import_id={import_id}' # Backward compat or hints
        })

    except Exception as e:
        current_app.logger.error(f"Error importing Excel: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@main_bp.route('/search', methods=['GET'])
def search():
    """ Handles drug name search with pagination. Returns JSON. """
    drug_name = request.args.get('drug_name')
    batch_id_search = request.args.get('batch_id_search')
    import_id = request.args.get('import_id')

    if not drug_name and not batch_id_search and not import_id:
        return jsonify({'error': 'No search parameters provided'}), 400

    page = request.args.get('page', 1, type=int)
    view = request.args.get('view')
    
    # Limitation Logic: OpenFDA = 10, FDALabel/Internal = 100000
    from dashboard.services.fdalabel_db import FDALabelDBService
    if FDALabelDBService.check_connectivity() or import_id:
        limit = 100000
    else:
        limit = 10 # OpenFDA limit
        
    skip = (page - 1) * limit

    if import_id:
        import_path = os.path.join(Config.UPLOAD_FOLDER, f"import_{import_id}.json")
        if os.path.exists(import_path):
            with open(import_path, 'r', encoding='utf-8') as f:
                all_imported_labels = json.load(f)
            total = len(all_imported_labels)
            labels = all_imported_labels[skip : skip + limit]
            
            # Enrich with favorite status if user logged in
            if current_user.is_authenticated:
                set_ids = [l['set_id'] for l in labels]
                favs = Favorite.query.filter(Favorite.user_id == current_user.id, Favorite.set_id.in_(set_ids)).all()
                fav_ids = {f.set_id for f in favs}
                for l in labels:
                    l['is_favorite'] = l['set_id'] in fav_ids

            drug_name_display = "Excel Import"
            page_title = "Imported FDALabel Results"
        else:
            return jsonify({'error': "Import session expired or not found."}), 404
    elif batch_id_search:
        ids_list = [sid.strip() for sid in batch_id_search.split(',') if sid.strip()]
        labels, total = find_labels_by_set_ids(ids_list, skip=skip, limit=limit)
        drug_name_display = "Identifier Batch Search"
        page_title = "Batch Search Results"
    else:
        labels, total = find_labels(drug_name, skip=skip, limit=limit)
        drug_name_display = drug_name
        page_title = f"Search Results for \"{drug_name}\""

    # Determine default view if not specified
    if not view:
        if total > 5 or batch_id_search or import_id:
             view = 'table'
        else:
             view = 'panel'

    is_internal = FDALabelDBService.check_connectivity()
    
    return jsonify({
        "drug_name": drug_name_display,
        "page_title": page_title,
        "labels": labels,
        "total": total,
        "page": page,
        "limit": limit,
        "view": view,
        "is_internal": is_internal,
        "batch_id_search": batch_id_search,
        "import_id": import_id
    })


@main_bp.route('/label/<set_id>')
def view_label(set_id):
    """ Returns the parsed SPL label data as JSON. """
    drug_name_from_query = request.args.get('drug_name', '')
    import_id = request.args.get('import_id')
    
    label_xml_raw = get_label_xml(set_id)
    if not label_xml_raw:
        return jsonify({'error': "Could not fetch the label data."}), 500

    doc_title, sections, fallback_html, highlights, table_of_contents, product_data = parse_spl_xml(label_xml_raw, set_id)
    
    metadata = extract_metadata_from_xml(label_xml_raw)
    
    # Re-use the clean_name logic for the display drug name
    def clean_header_text(text):
        if not text: return text
        if "highlights do not include" in text.lower():
            # Try to extract name between parentheses
            m = re.search(r'\(([^)]+)\)', text)
            if m:
                return m.group(1).strip()
            # Otherwise aggressively strip disclaimer
            text = re.sub(r'^These highlights do not include.*?safely and effectively\.\s*(See full prescribing information for.*?\.)?\s*', '', text, flags=re.IGNORECASE | re.DOTALL).strip()
        return text

    original_title = doc_title # Keep the raw title
    display_drug_name = drug_name_from_query if drug_name_from_query else clean_header_text(doc_title)

    # Load saved annotations
    saved_annotations = []
    from sqlalchemy import or_
    query = Annotation.query.filter(Annotation.set_id == set_id)
    if current_user.is_authenticated:
        query = query.filter(or_(Annotation.user_id == current_user.id, Annotation.is_public == True))
    else:
        query = query.filter(Annotation.is_public == True)
        
    user_annotations = query.all()
    for ann in user_annotations:
        saved_annotations.append({
            'id': str(ann.id),
            'section_number': ann.section_number,
            'question': ann.question,
            'answer': ann.answer,
            'keywords': json.loads(ann.keywords) if ann.keywords else [],
            'is_public': getattr(ann, 'is_public', False)
        })

    metadata = extract_metadata_from_xml(label_xml_raw)
    
    if not metadata or (metadata.get('brand_name') == 'Unknown Drug' and metadata.get('generic_name') == 'Unknown Generic'):
        ext_metadata = get_label_metadata(set_id, import_id=import_id)
        if ext_metadata:
            if not metadata: 
                metadata = ext_metadata
            else:
                for k, v in ext_metadata.items():
                    if not metadata.get(k) or metadata.get(k) in ['N/A', 'Unknown Drug', 'Unknown Generic', 'Unknown Manufacturer']:
                        metadata[k] = v

    brand_name = metadata.get('brand_name') if metadata and metadata.get('brand_name') != 'N/A' else None
    generic_name = metadata.get('generic_name') if metadata and metadata.get('generic_name') != 'N/A' else None
    manufacturer_name = metadata.get('manufacturer_name') if metadata and metadata.get('manufacturer_name') != 'N/A' else None
    effective_time = metadata.get('effective_time') if metadata else ''
    label_format = metadata.get('label_format') if metadata else None
    ndc = metadata.get('ndc', 'N/A') if metadata else 'N/A'
    application_number = metadata.get('application_number', 'N/A') if metadata else 'N/A'
    version_number = metadata.get('version_number') if metadata else None
    document_type = metadata.get('document_type') if metadata else None
    has_boxed_warning = metadata.get('has_boxed_warning') if metadata else False
    
    clean_app_num = None
    if application_number and application_number != 'N/A':
        first_app = application_number.split(',')[0].strip()
        clean_app_num = ''.join(filter(str.isdigit, first_app))

    # Simplified display_drug_name logic: just pick the best single name available
    # The frontend will handle combining Brand - Generic - Date
    if brand_name and brand_name != 'Unknown Drug':
        display_drug_name = brand_name
    elif generic_name and generic_name != 'Unknown Generic':
        display_drug_name = generic_name
    else:
        # doc_title was already cleaned by parse_spl_xml
        display_drug_name = display_drug_name 

    if metadata and metadata.get('faers_search_name') and metadata.get('faers_search_name') != 'Unknown Generic':
        faers_drug_name = metadata.get('faers_search_name')
    else:
        faers_drug_name = generic_name if generic_name and generic_name != 'Unknown Generic' else brand_name
    
    if not faers_drug_name or faers_drug_name == 'Unknown Drug' or faers_drug_name == 'Unknown Generic':
        temp_name = drug_name_from_query if drug_name_from_query else original_title
        if temp_name and len(temp_name) > 80: 
             paren_match = re.search(r'\(([^)]+)\)', temp_name)
             if paren_match:
                 faers_drug_name = paren_match.group(1).strip()
             else:
                 faers_drug_name = temp_name[:50]
        else:
            faers_drug_name = temp_name

    if faers_drug_name and faers_drug_name != 'N/A':
        faers_drug_name = re.split(r'[,;]', faers_drug_name)[0].strip()
        faers_drug_name = re.sub(r'\d+(\.\d+)?\s*(mg|mcg|g|ml|%|unit|iu)\b.*$', '', faers_drug_name, flags=re.IGNORECASE).strip()
        faers_drug_name = re.sub(r'\s+(tablet|capsule|injection|cream|ointment|gel|solution|suspension|spray|inhaler|powder).*$', '', faers_drug_name, flags=re.IGNORECASE).strip()

    tox_summary = {'dili': False, 'dict': False, 'diri': False}
    try:
        tox_agent = ToxAgent.query.filter_by(set_id=set_id, current='Yes').first()
        if tox_agent:
            tox_summary['dili'] = bool(tox_agent.dili_report and '<div' in tox_agent.dili_report)
            tox_summary['dict'] = bool(tox_agent.dict_report and '<div' in tox_agent.dict_report)
            tox_summary['diri'] = bool(tox_agent.diri_report and '<div' in tox_agent.diri_report)
            tox_summary['last_updated'] = tox_agent.last_updated.isoformat()
    except Exception as e:
        current_app.logger.error(f"Error fetching tox summary: {e}")

    return jsonify({
        'drug_name': display_drug_name,
        'brand_name': brand_name,
        'generic_name': generic_name,
        'original_title': original_title,
        'faers_drug_name': faers_drug_name,
        'manufacturer_name': manufacturer_name or '',
        'effective_time': effective_time,
        'label_format': label_format,
        'ndc': ndc,
        'application_number': application_number,
        'version_number': version_number,
        'document_type': document_type,
        'has_boxed_warning': has_boxed_warning,
        'clean_app_num': clean_app_num,
        'original_search': drug_name_from_query,
        'sections': sections, 
        'fallback_html': fallback_html, 
        'highlights': highlights, 
        'table_of_contents': table_of_contents,
        'product_data': product_data,
        'label_xml_raw': label_xml_raw,
        'set_id': set_id,
        'metadata': metadata,
        'saved_annotations': saved_annotations,
        'tox_summary': tox_summary,
        'user_id': current_user.id if current_user.is_authenticated else None
    })

@main_bp.route('/preferences', methods=['POST'])
@login_required
def preferences():
    if request.is_json:
        data = request.get_json()
    else:
        data = request.form

    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400

    if 'ai_provider' in data:
        current_user.ai_provider = data.get('ai_provider')
    if 'custom_gemini_key' in data:
        current_user.custom_gemini_key = data.get('custom_gemini_key')
    if 'openai_api_key' in data:
        current_user.openai_api_key = data.get('openai_api_key')
    if 'openai_base_url' in data:
        current_user.openai_base_url = data.get('openai_base_url')
    if 'openai_model_name' in data:
        current_user.openai_model_name = data.get('openai_model_name')
    
    db.session.commit()
    return jsonify({'success': True, 'message': 'Preferences saved successfully!'})

@main_bp.route("/snippet-preview", methods=["GET"])
def snippet_preview():
    drug_name = request.args.get("drug_name", "").strip()
    if not drug_name:
        return jsonify({"error": "Missing drug_name"}), 400

    try:
        drug_info = FDALabelDBService.get_drug_info(drug_name)
        if drug_info is None:
            return jsonify({"found": False, "drug_name": drug_name}), 200

        data = {
            "found": True,
            "set_id": drug_info["set_id"],
            "appr_num": drug_info["appr_num"],
            "product_name": drug_info["product_name"],
            "generic_name": drug_info["generic_name"],
            "active_ingredients": drug_info["active_ingredients"],
            "is_RLD": drug_info["is_RLD"],
            "effective_date": drug_info["effective_date"]
        }
        return jsonify(data)
    except Exception as e:
        current_app.logger.error(f"Error in snippet_preview: {e}")
        return jsonify({"error": str(e)}), 500

## export excel of a selected project.
def _safe_str(v):
    if v is None: return ''
    return str(v)

def _fmt_eff_time(v):
    if v is None: return ''
    if isinstance(v, datetime): return v.strftime('%Y/%m/%d')
    s = str(v).strip()
    if not s: return ''
    digits = ''.join([c for c in s if c.isdigit()])
    if len(digits) == 8:
        y, m, d = digits[:4], digits[4:6], digits[6:8]
        return f'{y}/{m}/{d}'
    try:
        dt = datetime.fromisoformat(s.replace('Z', ''))
        return dt.strftime('%Y/%m/%d')
    except Exception:
        return s

@main_bp.route('/export_project', methods=['GET'])
@login_required
def export_project():
    project_id = request.args.get('project_id', type=int)
    if not project_id:
        return jsonify({'success': False, 'error': 'Missing project_id'}), 400

    try:
        favs = Favorite.query.filter_by(user_id=current_user.id, project_id=project_id).all()
        if not favs:
            return jsonify({'success': False, 'error': 'No labels found for this project'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "FDALabel Export"
        headers = [
            "SET ID", "Trade Name", "Generic/Proper Name(s)", "Company",
            "SPL Effective Date (YYYY/MM/DD)", "Application Number(s)",
            "Labeling Type", "NDC(s)", "Marketing Category", "Dosage Form(s)",
            "Route(s) of Administration", "Established Pharmacologic Class(es)",
            "Active Ingredient(s)", "FDALabel Link", "DailyMed SPL Link", "DailyMed PDF Link",
            "Product Type", "Label Format", "Source", "Project ID"
        ]
        ws.append(headers)

        for f in favs:
            ws.append([
                _safe_str(getattr(f, 'set_id', None)),
                _safe_str(getattr(f, 'brand_name', None)),
                _safe_str(getattr(f, 'generic_name', None)),
                _safe_str(getattr(f, 'manufacturer_name', None)),
                _fmt_eff_time(getattr(f, 'effective_time', None)),
                _safe_str(getattr(f, 'application_number', None)),
                _safe_str(getattr(f, 'labeling_type', None)),
                _safe_str(getattr(f, 'ndc', None)),
                _safe_str(getattr(f, 'marketing_category', None)),
                _safe_str(getattr(f, 'dosage_forms', None)),
                _safe_str(getattr(f, 'routes', None)),
                _safe_str(getattr(f, 'epc', None)),
                _safe_str(getattr(f, 'active_ingredients', None)),
                _safe_str(getattr(f, 'fdalabel_link', None)),
                _safe_str(getattr(f, 'dailymed_spl_link', None)),
                _safe_str(getattr(f, 'dailymed_pdf_link', None)),
                _safe_str(getattr(f, 'product_type', None)),
                _safe_str(getattr(f, 'label_format', None)),
                _safe_str(getattr(f, 'source', None)),
                str(project_id),
            ])

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        widths = {"A": 40, "B": 28, "C": 32, "D": 28, "E": 18, "F": 22, "G": 18, "H": 24, "I": 22, "J": 22, "K": 22, "L": 30, "M": 36, "N": 36, "O": 36, "P": 36}
        for col, w in widths.items(): ws.column_dimensions[col].width = w

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"project_{project_id}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        current_app.logger.exception(f"Error exporting project {project_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

def _norm_str(v):
    if v is None: return ""
    return str(v).strip()

def _parse_eff_time_to_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    s = str(v).strip()
    if not s or s.lower() == "n/a": return None
    digits = "".join([c for c in s if c.isdigit()])
    if len(digits) == 8:
        try: return datetime(int(digits[0:4]), int(digits[4:6]), int(digits[6:8])).date()
        except Exception: return None
    try: return datetime.fromisoformat(s.replace("Z", "")).date()
    except Exception: return None

@main_bp.route('/project_stats', methods=['GET'])
@login_required
def project_stats():
    project_id = request.args.get('project_id', type=int)
    ingredient = request.args.get('ingredient', default=None, type=str)
    if not project_id:
        return jsonify({"success": False, "error": "Missing project_id"}), 400

    try:
        favs = Favorite.query.filter_by(user_id=current_user.id, project_id=project_id).all()
        if not favs:
            return jsonify({
                "success": True, "project_id": project_id, "total_labels": 0, "unique_manufacturers": 0, "unique_brands": 0,
                "date_min": None, "date_max": None, "product_type_counts": {}, "top_manufacturers": [],
                "document_type": {"raw": {}, "buckets": {"human_rx": 0, "human_otc": 0, "vaccine": 0, "animal_rx": 0, "animal_otc": 0, "other": 0, "unknown": 0}, "note": "No labels in project."},
                "effective_time_summary": {"dated_labels": 0, "missing_effective_time": 0, "filled_from_db": 0},
                "cumulative_by_effective_time": [], "ingredient_breakdown": None
            }), 200

        manufacturers, brands, eff_dates, product_types, set_ids, missing_date_set_ids = [], [], [], [], [], []
        for f in favs:
            sid = str(f.set_id) if f.set_id else None
            if sid: set_ids.append(sid)
            manufacturers.append(_norm_str(getattr(f, 'manufacturer_name', None)) or "Unknown")
            brands.append(_norm_str(getattr(f, 'brand_name', None)) or "Unknown")
            d = _parse_eff_time_to_date(getattr(f, 'effective_time', None))
            if d: eff_dates.append(d)
            elif sid: missing_date_set_ids.append(sid)
            pt = _norm_str(getattr(f, 'product_type', None))
            if not pt:
                lt = _norm_str(getattr(f, 'labeling_type', None)).upper()
                mc = _norm_str(getattr(f, 'market_category', None)).upper()
                pt = "OTC" if "OTC" in lt or "OTC" in mc else ("Rx" if lt or mc else "Unknown")
            product_types.append(pt)

        manu_set = {m for m in manufacturers if m and m != "Unknown"}
        brand_set = {b for b in brands if b and b != "Unknown"}
        internal_ok = FDALabelDBService.check_connectivity()
        filled_from_db = 0
        if internal_ok and missing_date_set_ids:
            db_map = FDALabelDBService.effective_time_map_for_set_ids(missing_date_set_ids)
            for sid in missing_date_set_ids:
                d = _parse_eff_time_to_date(db_map.get(sid))
                if d:
                    eff_dates.append(d)
                    filled_from_db += 1

        date_min = min(eff_dates).isoformat() if eff_dates else None
        date_max = max(eff_dates).isoformat() if eff_dates else None
        pt_counts = dict(Counter(product_types))
        manu_counts = Counter(manufacturers)
        top_manufacturers = [{"name": name, "count": count} for name, count in manu_counts.most_common(10) if name and name != "Unknown"]

        if internal_ok and set_ids:
            document_type = FDALabelDBService.document_type_breakdown_for_set_ids(set_ids)
        else:
            document_type = {"raw": {}, "buckets": {"human_rx": 0, "human_otc": 0, "vaccine": 0, "animal_rx": 0, "animal_otc": 0, "other": 0, "unknown": len(set_ids) or len(favs)}, "note": "Internal DB not available."}

        daily = Counter(eff_dates)
        cumulative_by_effective_time, cum = [], 0
        for day in sorted(daily.keys()):
            cum += daily[day]
            cumulative_by_effective_time.append({"date": day.isoformat(), "cumulative_count": cum})

        ingredient_breakdown = None
        if ingredient and ingredient.strip():
            q = ingredient.strip()
            if internal_ok and set_ids:
                ingredient_breakdown = FDALabelDBService.ingredient_role_breakdown_for_set_ids(set_ids=set_ids, substance_name=q)
            else:
                ingredient_breakdown = {"query": q, "active_count": 0, "inactive_count": 0, "both_count": 0, "not_found_count": len(set_ids) or len(favs), "matches": {}, "note": "Internal DB not available."}

        return jsonify({
            "success": True, "project_id": project_id, "total_labels": len(favs), "unique_manufacturers": len(manu_set), "unique_brands": len(brand_set),
            "date_min": date_min, "date_max": date_max, "product_type_counts": pt_counts, "top_manufacturers": top_manufacturers, "document_type": document_type,
            "effective_time_summary": {"dated_labels": len(eff_dates), "missing_effective_time": max(len(favs) - len(eff_dates), 0), "filled_from_db": filled_from_db},
            "cumulative_by_effective_time": cumulative_by_effective_time, "ingredient_breakdown": ingredient_breakdown
        }), 200
    except Exception as e:
        current_app.logger.exception(f"Error computing project_stats for {project_id}: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
