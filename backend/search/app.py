from flask import Flask, request, jsonify, Response, stream_with_context, send_file
from flask_cors import CORS

import os
import io
import json
import logging
import oracledb
import time
import threading
import pandas as pd

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from search.call_llm import safe_llm_call
from search.prompt_active import prompt_boring

# [OK] Import refactored logic from search.scripts/search_v1.py
from search.scripts.search_v1 import (
    search_v1 as search_v1_func,
    generate_answer_stream as generate_answer_stream_func,
    get_metadata as get_metadata_func,
    call_llm_stream as call_llm_stream_func,
    lob_to_string,  # used by export_xml
)
from search.scripts.search_v2 import search_v2 as search_v2_func


from search.scripts.search_v2 import (
    AgentState,
    run_controller,
    build_debug_stats,
    convert_oracle_to_filtered_results,
)

from search.scripts.search_v2_core.agents.answer_composer import build_answer_messages, run_answer_composer
from search.scripts.search_v2_core.agents.reasoning_generator import run_reasoning_generator

from search.scripts.search_v2_core.config import client
try:
    from search.scripts.search_v2_core.config import llm_model_name  # optional
except Exception:
    llm_model_name = None

# Initialize Flask app
app = Flask(__name__)
CORS(app, origins=["http://localhost:8845", "http://ncshpcgpu01:8845", "https://elsa.fda.gov"])

load_dotenv()
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)
logging.getLogger("openai").setLevel(logging.WARNING)
logging.getLogger("httpx").setLevel(logging.WARNING)

# ---- LLM init (still needed for random_query, and potentially others) ----
openai_api_key = os.getenv("LLM_KEY", "")
openai_api_base = os.getenv("LLM_URL", "")
llm_model_name = os.getenv("LLM_MODEL", "")

client = OpenAI(api_key=openai_api_key, base_url=openai_api_base)

# ---- DB init (still needed for export_xml/export_excel) ----
FDALabel_SERV = os.getenv("FDALabel_SERV")
FDALabel_PORT = os.getenv("FDALabel_PORT")
FDALabel_APP = os.getenv("FDALabel_APP")
FDALabel_USER = os.getenv("FDALabel_USER")
FDALabel_PSW = os.getenv("FDALabel_PSW")

dsnStr = oracledb.makedsn(FDALabel_SERV, FDALabel_PORT, FDALabel_APP)


# -------------------------
# Thin wrappers over search_v1
# -------------------------
@app.route("/api/search", methods=["POST"])
def search():
    payload = request.json or {}
    resp, status = search_v1_func(payload)
    return jsonify(resp), status


@app.route("/api/search_agentic", methods=["POST"])
def search_agentic():
    payload = request.json or {}
    resp, status = search_v2_func(payload)
    return jsonify(resp), status


@app.route("/api/generate_answer", methods=["POST"])
def generate_answer():
    payload = request.json or {}

    def gen():
        yield from generate_answer_stream_func(payload)

    return Response(stream_with_context(gen()), content_type="text/plain")


@app.route("/api/get_metadata", methods=["POST"])
def get_metadata():
    payload = request.json or {}
    resp, status = get_metadata_func(payload)
    return jsonify(resp), status


def build_v2_response_from_state(state: AgentState):
    debug_stats = build_debug_stats(state)

    raw_results = state.retrieval.get("results", [])
    processed_results_dict = convert_oracle_to_filtered_results(raw_results)
    final_results = list(processed_results_dict.values())

    response = {
        "med_answer": state.answer["response_text"],
        "debug_intent": state.intent,
        "results": final_results,
        "is_answerable": True,
        "input_type": "T1",
        "generated_sql": state.retrieval.get("generated_sql", ""),
        "total_counts": len(final_results),
        "suggestions": [],
        "agent_flow": state.agent_flow,
        "reasoning": state.reasoning,
        "debug_plan": state.retrieval.get("plan", {}),
        "debug_stats": debug_stats,
        "trace_log": state.trace_log,
    }
    return response, 200

def _humanize_trace(line: str) -> str:
    # Optional: convert noisy trace lines into friendly step labels
    s = (line or "").strip()
    if s.startswith("Planner:"):
        return "Planning query strategy..."
    if "db_executor" in s.lower() or s.startswith("DB"):
        return "Running database query..."
    if "Evidence Fetcher" in s or "evidence" in s.lower():
        return "Fetching label evidence..."
    if "answer_composer" in s.lower() or "Composer" in s:
        return "Composing answer..."
    # fallback: show raw trace line
    return s

def stream_answer_tokens(state):
    """
    Yields answer text chunks (token deltas) from the LLM.
    Falls back to non-streamed answer for deterministic paths.
    """
    intent_type = (state.intent or {}).get("type") or ""
    is_aggregate = bool((state.retrieval or {}).get("aggregate"))

    # Deterministic/simple paths: just run composer and yield once
    if is_aggregate or intent_type in ("chitchat", "clarification"):
        run_answer_composer(state)
        yield (state.answer or {}).get("response_text", "") or ""
        return

    # Build messages the same way answer_composer does
    messages = build_answer_messages(state)

    # IMPORTANT: set your actual model if MODEL_NAME isn't available
    model = llm_model_name or ""  # <-- change if needed

    # Stream tokens
    stream = client.chat.completions.create(
        model=model,
        messages=messages,
        temperature=0.1,
        stream=True,
    )

    for evt in stream:
        try:
            delta = evt.choices[0].delta.content
        except Exception:
            delta = None
        if delta:
            yield delta


@app.route("/api/search_agentic_stream", methods=["POST"])
def search_agentic_stream():
    payload = request.json or {}

    def generate():
        state = AgentState(payload)
        done = threading.Event()
        err = {}

        def worker():
            try:
                # [OK] stop BEFORE answer generation so we can stream tokens ourselves
                run_controller(state, stop_before="answer_composer")
            except Exception as e:
                err["e"] = e
            finally:
                done.set()

        threading.Thread(target=worker, daemon=True).start()

        sent = 0
        yield json.dumps({"type": "status", "text": "Starting agent run..."}) + "\n"

        # Stream trace/status while retrieval is running
        while not done.is_set():
            while sent < len(state.trace_log):
                line = state.trace_log[sent]
                yield json.dumps({"type": "status", "text": _humanize_trace(line)}) + "\n"
                sent += 1
            time.sleep(0.12)

        # Flush remaining trace lines
        while sent < len(state.trace_log):
            line = state.trace_log[sent]
            yield json.dumps({"type": "status", "text": _humanize_trace(line)}) + "\n"
            sent += 1

        if "e" in err:
            yield json.dumps({"type": "error", "error": str(err["e"])}) + "\n"
            return

        # [OK] Stream answer tokens
        yield json.dumps({"type": "status", "text": "Writing answer..."}) + "\n"
        yield json.dumps({"type": "answer_start"}) + "\n"

        answer_text = ""
        try:
            for tok in stream_answer_tokens(state):
                if not tok:
                    continue
                answer_text += tok
                yield json.dumps({"type": "chunk", "text": tok}) + "\n"
        except Exception as e:
            yield json.dumps({"type": "error", "error": f"Answer streaming failed: {str(e)}"}) + "\n"
            return

        state.answer = state.answer or {}
        state.answer["response_text"] = answer_text

        yield json.dumps({"type": "answer_end"}) + "\n"

        # [OK] Now generate reasoning AFTER we have the final answer text
        try:
            yield json.dumps({"type": "status", "text": "Finalizing reasoning..."}) + "\n"
            run_reasoning_generator(state)
        except Exception:
            # Don't fail the whole stream if reasoning fails
            pass

        resp, status = build_v2_response_from_state(state)
        yield json.dumps({"type": "final", "status": status, "payload": resp}) + "\n"

    r = Response(stream_with_context(generate()), mimetype="application/x-ndjson")
    r.headers["Cache-Control"] = "no-cache"
    r.headers["X-Accel-Buffering"] = "no"  # nginx
    return r

# -------------------------
# Chat endpoint (wrapper)
# -------------------------
@app.route("/api/chat", methods=["POST"])
def chat():
    try:
        chatHistory = json.loads(request.form.get("chatHistory", "[]"))
        documents = request.form.get("documents", "")
        doc_type = request.form.get("doc_type", "")

        uploaded_files = []
        if "uploadedFiles" in request.files:
            uploaded_files = [f for f in request.files.getlist("uploadedFiles") if f and f.filename]

        spl_xmls = []
        if doc_type == "setids":
            # NOTE: app.py used get_xml() previously but it wasn't shown in your snippet.
            # If get_xml is defined elsewhere, keep importing it and use it here.
            # Otherwise, switch your frontend to send 'ref' (raw XML) or move get_xml logic into search_v1.
            setids = json.loads(documents)
            from search.file_util import get_xml  # keep local import to reduce global coupling
            spl_xmls = get_xml(setids)
        elif doc_type == "ref":
            spl_xmls.append(documents)
        else:
            return jsonify({"error": "Invalid doc_type"}), 400

        def chat_stream():
            # search_v1's call_llm_stream returns JSON-lines streaming chunks
            yield from search.call_llm_stream_func(chatHistory, spl_xmls, uploaded_files)

        return Response(stream_with_context(chat_stream()), content_type="application/json")

    except Exception as e:
        logger.error(f"Error in chat: {e}")
        return jsonify({"error": str(e)}), 500


# -------------------------
# Export XML (still local)
# -------------------------
@app.route("/api/export_xml", methods=["POST"])
def export_xml():
    """
    Receives a list of SET_IDs, fetches their corresponding SPL_XML from the Oracle database,
    and returns a JSON object mapping each SET_ID to its XML content.
    """
    try:
        data = request.json or {}
        set_ids = data.get("set_ids", [])

        if len(set_ids) > 5:
            set_ids = set_ids[:5]

        if not set_ids:
            return jsonify({"error": "No SET_IDs provided"}), 400

        xml_data = {}
        con = None
        try:
            con = oracledb.connect(user=FDALabel_USER, password=FDALabel_PSW, dsn=dsnStr)
            cursor = con.cursor()
            logger.info("Database connection established for XML export.")

            query = "SELECT XMLSERIALIZE(DOCUMENT spl_xml AS CLOB) FROM spl WHERE set_id = :set_id"

            for set_id in set_ids:
                cursor.execute(query, {"set_id": set_id})
                result = cursor.fetchone()
                if result:
                    xml_data[set_id] = lob_to_string(result[0])
                else:
                    xml_data[set_id] = "XML content not found in database."

        except oracledb.DatabaseError as db_err:
            logger.error(f"Database error during XML export: {db_err}")
            return jsonify({"error": "Database operation failed"}), 500
        finally:
            if con:
                con.close()

        return jsonify(xml_data)

    except Exception as e:
        logger.error(f"Error in /api/export_xml: {e}")
        return jsonify({"error": str(e)}), 500


# -------------------------
# Export Excel (still local)
# -------------------------
@app.route("/api/export_excel", methods=["POST"])
def export_excel():
    try:
        data = request.json or {}
        set_ids = data.get("set_ids", [])

        if not set_ids:
            return jsonify({"error": "No SET_IDs provided"}), 400

        con = None
        try:
            con = oracledb.connect(user=FDALabel_USER, password=FDALabel_PSW, dsn=dsnStr)
            cursor = con.cursor()

            bind_names = [f":id{i}" for i in range(len(set_ids))]
            bind_map = {f"id{i}": sid for i, sid in enumerate(set_ids)}

            sql = f"SELECT * FROM druglabel.dgv_sum_rx_spl WHERE set_id IN ({','.join(bind_names)})"
            cursor.execute(sql, bind_map)

            columns = [col[0].upper() for col in cursor.description]
            rows = cursor.fetchall()
            df_db = pd.DataFrame(rows, columns=columns)

        except Exception as e:
            logger.error(f"DB Error in export_excel: {e}")
            return jsonify({"error": "Database query failed"}), 500
        finally:
            if con:
                con.close()

        def get_col(df, candidates):
            for c in candidates:
                if c in df.columns:
                    return df[c]
            return ""

        export_df = pd.DataFrame()
        export_df["Marketing Category"] = get_col(df_db, ["MARKET_CATEGORIES", "MARKETING_CATEGORIES", "APPLICATION_TYPE"])
        export_df["Application Number(s)"] = get_col(df_db, ["APPR_NUM", "APPROVAL_NUM"])
        export_df["Trade Name"] = get_col(df_db, ["PRODUCT_NAMES", "PRODUCT_TITLE"])
        export_df["Generic/Proper Name(s)"] = get_col(df_db, ["GENERIC_NAMES", "PRODUCT_NORMD_GENERIC_NAMES"])

        eff_time = get_col(df_db, ["EFF_TIME", "EFFECTIVE_TIME"])
        export_df["SPL Effective Date (YYYY/MM/DD)"] = pd.to_datetime(eff_time, errors="coerce").dt.strftime("%Y/%m/%d")

        export_df["Initial U.S. Approval"] = get_col(df_db, ["INITIAL_APPROVAL_YEAR", "APPROVAL_YEAR"])
        export_df["Dosage Form(s)"] = get_col(df_db, ["DOSAGE_FORMS"])
        export_df["Route(s) of Administration"] = get_col(df_db, ["ROUTES", "ROUTES_OF_ADMINISTRATION"])
        export_df["Established Pharmacologic Class(es)"] = get_col(df_db, ["EPC"])
        export_df["Company"] = get_col(df_db, ["AUTHOR_ORG_NORMD_NAME", "COMPANY"])
        export_df["NDC(s)"] = get_col(df_db, ["NDC_CODES", "NDC"])

        export_df["Marketing Date(s) (YYYY/MM/DD)"] = get_col(df_db, ["MARKETING_START_DATE", "START_MARKETING_DATE"])
        export_df["Active Ingredient UNII(s)"] = get_col(df_db, ["ACT_INGR_UNIIS", "UNII"])
        export_df["Active Ingredient(s)"] = get_col(df_db, ["ACT_INGR_NAMES", "ACTIVE_INGREDIENTS"])
        export_df["Labeling Type"] = get_col(df_db, ["DOCUMENT_TYPE", "DOC_TYPE", "LABEL_TYPE"])

        base_fda = "https://fdalabel.fda.gov/fdalabel-r/services/spl/set-ids/{}/spl-doc"
        base_dm_spl = "https://dailymed.nlm.nih.gov/dailymed/lookup.cfm?setid={}"
        base_dm_pdf = "https://dailymed.nlm.nih.gov/dailymed/downloadpdffile.cfm?setId={}"

        set_ids_col = get_col(df_db, ["SET_ID"])
        export_df["FDALabel Link"] = set_ids_col.apply(lambda x: base_fda.format(x) if x else "")
        export_df["DailyMed SPL Link"] = set_ids_col.apply(lambda x: base_dm_spl.format(x) if x else "")
        export_df["DailyMed PDF Link"] = set_ids_col.apply(lambda x: base_dm_pdf.format(x) if x else "")
        export_df["SET ID"] = set_ids_col

        template_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            "ideas",
            "fdalabel-query-240623-DILI-RLD-WARNING.xlsx",
        )

        try:
            wb = load_workbook(template_path)
            ws = wb.active

            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            for r in dataframe_to_rows(export_df, index=False, header=False):
                ws.append(r)

            out_buffer = io.BytesIO()
            wb.save(out_buffer)
            out_buffer.seek(0)

            return send_file(
                out_buffer,
                as_attachment=True,
                download_name="fdalabel_export.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            logger.error(f"Template processing error: {e}")
            return jsonify({"error": f"Failed to generate Excel: {str(e)}"}), 500

    except Exception as e:
        logger.error(f"Error in export_excel: {e}")
        return jsonify({"error": str(e)}), 500


# -------------------------
# Random query (still local)
# -------------------------
@app.route("/api/random_query", methods=["GET"])
def random_query():
    try:
        messages = [{"role": "system", "content": prompt_boring}]
        process_success, generated_question = safe_llm_call(
            client, messages, max_tokens=100, temperature=0.9
        )
        if not process_success:
            return jsonify({"query": "What are the indications for Ozempic?"})

        return jsonify({"query": generated_question.replace('"', "").strip()})

    except Exception as e:
        logger.error(f"Error in random_query: {e}")
        return jsonify({"query": "What are the indications for Ozempic?"}), 500


@app.route("/api/snippet-preview", methods=["GET"])
def snippet_preview():
    """
    Returns basic info (NDA, Set ID, etc.) for a given drug name to be shown in a tooltip.
    Matches against PRODUCT_NAMES or PRODUCT_NORMD_GENERIC_NAMES.
    Sorts by EFFECTIVE_TIME DESC and returns the top 1.
    """
    drug_name = request.args.get("drug_name", "").strip()
    if not drug_name:
        return jsonify({"error": "Missing drug_name"}), 400

    con = None
    try:
        con = oracledb.connect(user=FDALabel_USER, password=FDALabel_PSW, dsn=dsnStr)
        cursor = con.cursor()

        # We use a simple pattern match. 
        # Note: DGV_SUM_SPL is usually in the DRUGLABEL schema, but might be aliased.
        # We'll assume the user connected with a user that can see it (druglabel.dgv_sum_spl or synonym).
        
        # We want to match whole words or start of string preferably, but for a snippet tool,
        # a simple LIKE %term% is often what's expected, though exact match is better for "Preview".
        # Let's try exact match or close match. 
        # Since the snippet logic found "Ibuprofen", we expect "Ibuprofen" to be in the name.
        
        query = """
            SELECT 
                s.SET_ID, 
                s.APPR_NUM, 
                s.PRODUCT_NAMES, 
                s.PRODUCT_NORMD_GENERIC_NAMES, 
                s.ACT_INGR_NAMES,
                rld.RLD,
                s.EFF_TIME
            FROM druglabel.DGV_SUM_SPL s
            LEFT JOIN druglabel.sum_spl_rld rld on rld.spl_id = s.spl_id
            WHERE (UPPER(s.PRODUCT_NAMES) LIKE UPPER(:dn) OR UPPER(s.PRODUCT_NORMD_GENERIC_NAMES) LIKE UPPER(:dn) OR UPPER(s.ACT_INGR_NAMES) LIKE UPPER(:dn))
            ORDER BY rld.RLD DESC, s.EFF_TIME DESC
            FETCH FIRST 1 ROWS ONLY
        """
        bind_val = f"{drug_name}"
        
        cursor.execute(query, {"dn": bind_val})
        row = cursor.fetchone()
        
        if not row: # try vague match
            bind_val = f"%{drug_name}%"
            cursor.execute(query, {"dn": bind_val})
            row = cursor.fetchone()    
            if not row: # if still no match, return False
                return jsonify({"found": False, "drug_name": drug_name}), 200

        # Map results
        # row: (SET_ID, SPL_ID, APPR_NUM, PROD_NAMES, GENERIC, ACT_INGR, EFF_DATE)
        data = {
            "found": True,
            "set_id": row[0],
            "appr_num": row[1],
            "product_name": row[2],
            "generic_name": row[3],
            "active_ingredients": row[4],
            "is_RLD": row[5],
            "effective_date": row[6]
        }
        return jsonify(data)

    except Exception as e:
        logger.error(f"Error in snippet_preview: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if con:
            con.close()


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", "5102"))
    app.run(host="0.0.0.0", port=port, debug=True)
